import cv2
import numpy as np
import os
import json
import csv
import re
import sys
import time
import math
import threading
from collections import deque
from datetime import datetime

# ============================================================
# EXCEL
# ============================================================
# openpyxl digunakan agar data inspeksi dapat ditambahkan ke file .xlsx
# yang sama setiap kali auto-capture atau capture manual dijalankan.
try:
    from openpyxl import Workbook as ExcelWorkbook
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    ExcelWorkbook = None
    load_workbook = None
    Alignment = Border = Font = PatternFill = Side = None
    get_column_letter = None
    OPENPYXL_AVAILABLE = False


# ============================================================
# KONFIGURASI UTAMA
# ============================================================
CAMERA_INDEX = 0
CAMERA_WIDTH = 640
CAMERA_HEIGHT = 480
CAMERA_FPS = 30
FLIP_CAMERA = True

MATCHING_THRESHOLD = 0.58
TEMPLATE_SCALES = (0.70, 0.80, 0.90, 1.00, 1.10, 1.20, 1.30)

# Deteksi harus muncul beberapa frame berturut-turut agar dianggap valid.
MIN_DETECTION_FRAMES = 4
# Deteksi yang hilang sesaat tetap ditahan agar tampilan tidak berkedip.
HOLD_MISSING_FRAMES = 6

# AUTO CAPTURE DUA TAHAP
# Tahap 1: PCB dipindai selama 5 detik sejak minimal satu komponen terbaca.
# Tahap 2: hasil GOOD/NOT GOOD harus stabil selama 5 detik sebelum capture.
INITIAL_SCAN_SECONDS = 5.0
STABLE_CAPTURE_SECONDS = 5.0
AUTO_CAPTURE_MIN_DETECTED_COMPONENTS = 1

# Jika semua komponen menghilang sesaat, proses tidak langsung dibatalkan.
AUTO_CAPTURE_EMPTY_CANCEL_FRAMES = 12

# Auto capture dibuka kembali setelah PCB diangkat/tidak terlihat beberapa frame.
AUTO_CAPTURE_UNLOCK_FRAMES = 15

# Pengaturan penyimpanan dataset pembelajaran.
DATASET_CROP_PADDING = 12
JPEG_QUALITY = 95

# Batas inspeksi kemiringan:
# - GOOD     : nilai absolut kemiringan < 15 derajat
# - NOT GOOD : nilai absolut kemiringan >= 15 derajat
# Nilai tepat 15 derajat dimasukkan ke NOT GOOD agar tidak ada area keputusan kosong.
# Batas pergeseran tetap menyesuaikan ukuran komponen.
ANGLE_TOLERANCE_DEG = 15.0
MIN_SHIFT_TOLERANCE_PX = 8.0
SHIFT_TOLERANCE_RATIO = 0.18

# Batas kestabilan sebelum auto capture.
MAX_POSITION_STD_PX = 5.0
MAX_ANGLE_STD_DEG = 4.0

PANEL_WIDTH = 500
MAX_PANEL_COMPONENTS = 9

# C1 dan C2 berbentuk bulat sehingga sudut kontur mudah berubah karena
# pantulan cahaya. Sudutnya diabaikan, tetapi kelengkapan dan pergeseran
# tetap diperiksa.
ROUND_COMPONENTS = {
    "C1", "C2",
    "CAPACITOR1", "CAPACITOR2",
    "CAPASITOR1", "CAPASITOR2",
}


# Untuk suara alarm di Windows.
if sys.platform == "win32":
    import winsound
else:
    try:
        import pygame

        pygame.mixer.init()
    except Exception:
        pygame = None


# ============================================================
# FUNGSI BANTU
# ============================================================
def safe_name(text):
    """Membuat nama aman untuk file/folder."""
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", str(text)).strip("_")
    return cleaned or "komponen"


def is_round_component(component):
    """Mengenali nama C1/C2 walaupun memakai spasi atau underscore."""
    compact = re.sub(r"[^A-Z0-9]+", "", str(component).upper())
    return compact in ROUND_COMPONENTS


def normalize_axial_angle(angle):
    """Normalisasi sudut benda ke rentang -90 sampai < 90 derajat."""
    return ((float(angle) + 90.0) % 180.0) - 90.0


def axial_angle_difference(angle, reference):
    """Selisih sudut dengan periodisitas 180 derajat."""
    return normalize_axial_angle(float(angle) - float(reference))


def axial_mean(angles):
    """Rata-rata sudut untuk objek yang orientasinya berulang setiap 180 derajat."""
    if not angles:
        return 0.0
    values = np.radians(np.asarray(angles, dtype=np.float64) * 2.0)
    mean_sin = float(np.mean(np.sin(values)))
    mean_cos = float(np.mean(np.cos(values)))
    if abs(mean_sin) < 1e-9 and abs(mean_cos) < 1e-9:
        return normalize_axial_angle(angles[-1])
    return normalize_axial_angle(math.degrees(0.5 * math.atan2(mean_sin, mean_cos)))


def axial_std(angles):
    """Simpangan baku pendekatan untuk sudut periodik 180 derajat."""
    if len(angles) < 2:
        return 0.0
    center = axial_mean(angles)
    diffs = [axial_angle_difference(value, center) for value in angles]
    return float(np.std(diffs))


def classify_template(template_name):
    """Mengambil nama komponen dasar dan kondisi dari nama file template."""
    original = str(template_name).strip()
    lowered = original.lower()

    not_good_patterns = (
        r"not[\s_-]*good",
        r"\bnotgood\b",
        r"(^|[\s_-])ng($|[\s_-])",
        r"(^|[\s_-])(cacat|salah|miring|geser|hilang|rusak)($|[\s_-])",
    )
    is_not_good = any(re.search(pattern, lowered) for pattern in not_good_patterns)

    # Hapus penanda kondisi supaya GOOD/NOT GOOD untuk komponen yang sama
    # memiliki ID komponen yang sama.
    cleaned = re.sub(r"not[\s_-]*good", " ", lowered)
    cleaned = re.sub(r"(^|[\s_-])(good|ok|ng|cacat|salah|miring|geser|hilang|rusak)($|[\s_-])", " ", cleaned)
    cleaned = re.sub(r"[\s_-]+", "_", cleaned).strip("_")

    if not cleaned:
        cleaned = lowered.replace(" ", "_")

    component_id = cleaned.upper()
    condition = "NOT_GOOD" if is_not_good else "GOOD"
    return component_id, condition


def detection_iou(det_a, det_b):
    ax, ay, aw, ah = det_a["bbox"]
    bx, by, bw, bh = det_b["bbox"]

    left = max(ax, bx)
    top = max(ay, by)
    right = min(ax + aw, bx + bw)
    bottom = min(ay + ah, by + bh)

    inter_w = max(0, right - left)
    inter_h = max(0, bottom - top)
    intersection = inter_w * inter_h
    union = aw * ah + bw * bh - intersection
    return intersection / union if union > 0 else 0.0


# ============================================================
# ALARM
# ============================================================
class AlarmSystem:
    def __init__(self):
        self.alarm_active = False
        self.alarm_muted = False
        self.fault_present = False
        self.stop_alarm_flag = False
        self.alarm_thread = None

    def update_fault(self, fault_present):
        self.fault_present = bool(fault_present)
        if self.fault_present and not self.alarm_muted:
            self.start_alarm()
        else:
            self.stop_alarm(clear_fault=False)

    def start_alarm(self):
        if self.alarm_active or self.alarm_muted:
            return
        self.alarm_active = True
        self.stop_alarm_flag = False
        if self.alarm_thread is None or not self.alarm_thread.is_alive():
            self.alarm_thread = threading.Thread(target=self._alarm_loop, daemon=True)
            self.alarm_thread.start()

    def stop_alarm(self, clear_fault=True):
        self.alarm_active = False
        self.stop_alarm_flag = True
        if clear_fault:
            self.fault_present = False
        self._stop_sound()

    def toggle_mute(self):
        self.alarm_muted = not self.alarm_muted
        if self.alarm_muted:
            self.stop_alarm(clear_fault=False)
        elif self.fault_present:
            self.start_alarm()
        return self.alarm_muted

    def _alarm_loop(self):
        sound_on = False
        last_toggle = 0.0
        while self.alarm_active and not self.stop_alarm_flag:
            now = time.time()
            if now - last_toggle >= 1.0:
                sound_on = not sound_on
                last_toggle = now
                if sound_on:
                    self._play_sound()
                else:
                    self._stop_sound()
            time.sleep(0.05)
        self._stop_sound()

    def _play_sound(self):
        try:
            if sys.platform == "win32":
                winsound.Beep(1000, 180)
            elif pygame is not None:
                for sound_file in ("alarm.wav", "alarm.mp3", "beep.wav", "warning.wav"):
                    if os.path.exists(sound_file):
                        pygame.mixer.Sound(sound_file).play()
                        return
                print("\a", end="", flush=True)
        except Exception:
            pass

    def _stop_sound(self):
        try:
            if sys.platform != "win32" and pygame is not None:
                pygame.mixer.stop()
        except Exception:
            pass


# ============================================================
# SMOOTHING
# ============================================================
class AngleSmoother:
    def __init__(self, window_size=15):
        self.window_size = window_size
        self.history = {}
        self.stable = {}

    def update(self, component, angle):
        angle = normalize_axial_angle(angle)
        if component not in self.history:
            self.history[component] = deque(maxlen=self.window_size)
        self.history[component].append(angle)
        self.stable[component] = axial_mean(list(self.history[component]))
        return self.stable[component]

    def get(self, component, default=0.0):
        return float(self.stable.get(component, default))

    def get_std(self, component):
        values = list(self.history.get(component, []))
        return axial_std(values)

    def clear(self, component=None):
        if component is None:
            self.history.clear()
            self.stable.clear()
        else:
            self.history.pop(component, None)
            self.stable.pop(component, None)


class PositionSmoother:
    def __init__(self, window_size=12):
        self.window_size = window_size
        self.history = {}
        self.stable = {}

    def update(self, component, center):
        x, y = float(center[0]), float(center[1])
        if component not in self.history:
            self.history[component] = deque(maxlen=self.window_size)
        self.history[component].append((x, y))

        values = np.asarray(self.history[component], dtype=np.float32)
        median_x = float(np.median(values[:, 0]))
        median_y = float(np.median(values[:, 1]))

        # Sedikit exponential smoothing sesudah median agar tetap responsif.
        if component in self.stable:
            old_x, old_y = self.stable[component]
            alpha = 0.35
            median_x = alpha * median_x + (1.0 - alpha) * old_x
            median_y = alpha * median_y + (1.0 - alpha) * old_y

        self.stable[component] = (median_x, median_y)
        return self.stable[component]

    def get_std(self, component):
        values = list(self.history.get(component, []))
        if len(values) < 2:
            return 0.0
        data = np.asarray(values, dtype=np.float32)
        std_x = float(np.std(data[:, 0]))
        std_y = float(np.std(data[:, 1]))
        return math.hypot(std_x, std_y)

    def clear(self, component=None):
        if component is None:
            self.history.clear()
            self.stable.clear()
        else:
            self.history.pop(component, None)
            self.stable.pop(component, None)


# ============================================================
# DETEKTOR KOMPONEN
# ============================================================
class ComponentDetector:
    REFERENCE_VERSION = 2

    def __init__(self, base_dir, template_folder="gambar_input"):
        self.base_dir = os.path.abspath(base_dir)
        self.template_folder = os.path.join(self.base_dir, template_folder)

        # Setiap jig memiliki file referensi sendiri. Referensi aktif selalu
        # mengikuti jig yang dipilih dengan tombol 1 atau 2.
        self.active_jig = 1
        self.reference_files = {
            1: os.path.join(self.base_dir, "referensi_jig_1.json"),
            2: os.path.join(self.base_dir, "referensi_jig_2.json"),
        }
        self.legacy_reference_file = os.path.join(self.base_dir, "referensi_komponen.json")
        self.reference_file = self.reference_files[self.active_jig]
        self.capture_folder = os.path.join(self.base_dir, "hasil_capture")
        # Dataset dipisahkan dari laporan inspeksi agar mudah dipakai kembali
        # untuk klasifikasi atau proses labeling YOLO.
        self.dataset_folder = os.path.join(self.base_dir, "dataset_pembelajaran")

        # File Excel dipisahkan berdasarkan jig:
        # data_excel_inspeksi/JIG_1/Data_Inspeksi_JIG_1.xlsx
        # data_excel_inspeksi/JIG_2/Data_Inspeksi_JIG_2.xlsx
        self.excel_base_folder = os.path.join(
            self.base_dir,
            "data_excel_inspeksi",
        )

        self.matching_threshold = MATCHING_THRESHOLD
        self.templates = []
        self.expected_components = []
        self.reference = None

        self.angle_smoother = AngleSmoother(window_size=15)
        self.position_smoother = PositionSmoother(window_size=12)
        self.tracks = {}

        self.alarm = AlarmSystem()
        self.scroll_offset = 0

        self.auto_capture_counter = 0
        self.auto_capture_locked = False
        self.incomplete_counter = 0
        self.auto_capture_phase = "WAITING"
        self.auto_scan_started_at = None
        self.auto_stable_started_at = None
        self.auto_capture_candidate_signature = None
        self.auto_capture_candidate_decision = None
        self.auto_scan_remaining_seconds = INITIAL_SCAN_SECONDS
        self.auto_capture_remaining_seconds = STABLE_CAPTURE_SECONDS
        self.auto_capture_empty_counter = 0
        self.last_capture_message = ""
        self.last_capture_message_until = 0.0

        self.board_present_counter = 0

        self.load_templates()
        self.load_reference()

    # --------------------------------------------------------
    # TEMPLATE DAN REFERENSI
    # --------------------------------------------------------
    def load_templates(self):
        os.makedirs(self.template_folder, exist_ok=True)
        loaded = []

        for filename in sorted(os.listdir(self.template_folder)):
            if not filename.lower().endswith((".png", ".jpg", ".jpeg", ".bmp")):
                continue

            path = os.path.join(self.template_folder, filename)
            image = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
            if image is None:
                continue

            max_size = 220
            height, width = image.shape
            if max(height, width) > max_size:
                scale = max_size / float(max(height, width))
                image = cv2.resize(
                    image,
                    (max(1, int(width * scale)), max(1, int(height * scale))),
                    interpolation=cv2.INTER_AREA,
                )

            image = cv2.GaussianBlur(image, (3, 3), 0)
            template_name = os.path.splitext(filename)[0]
            component, condition = classify_template(template_name)

            loaded.append(
                {
                    "image": image,
                    "name": template_name,
                    "component": component,
                    "condition": condition,
                    "path": path,
                }
            )

        self.templates = loaded
        self.expected_components = sorted({item["component"] for item in loaded})

        # Reset track saat template berubah.
        self.tracks = {
            component: {"seen": 0, "missed": HOLD_MISSING_FRAMES + 1, "last": None}
            for component in self.expected_components
        }
        self.angle_smoother.clear()
        self.position_smoother.clear()
        self.reset_auto_capture_lock()

    def load_reference(self):
        """Memuat referensi milik jig yang sedang aktif."""
        self.reference = None
        self.reference_file = self.reference_files[self.active_jig]

        # Migrasi satu kali: referensi program V3 dianggap sebagai referensi JIG 1.
        if (
            self.active_jig == 1
            and not os.path.exists(self.reference_file)
            and os.path.exists(self.legacy_reference_file)
        ):
            try:
                os.replace(self.legacy_reference_file, self.reference_file)
                print(f"[MIGRASI] referensi_komponen.json dipindahkan menjadi {os.path.basename(self.reference_file)}")
            except OSError as exc:
                print(f"[PERINGATAN] Referensi lama gagal dimigrasikan: {exc}")

        if not os.path.exists(self.reference_file):
            print(f"[INFO] Referensi JIG {self.active_jig} belum ada. Letakkan PCB GOOD lalu tekan K.")
            return

        try:
            with open(self.reference_file, "r", encoding="utf-8") as file:
                data = json.load(file)

            if data.get("version") != self.REFERENCE_VERSION or not isinstance(data.get("components"), dict):
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup = self.reference_file + f".backup_lama_{timestamp}"
                os.replace(self.reference_file, backup)
                print(f"[PERINGATAN] Format referensi JIG {self.active_jig} tidak cocok dan dipindahkan ke: {backup}")
                print(f"[INFO] Kalibrasi ulang JIG {self.active_jig}: pasang PCB GOOD lengkap lalu tekan K.")
                return

            stored_jig = data.get("jig_id")
            if stored_jig is not None and int(stored_jig) != self.active_jig:
                print(f"[PERINGATAN] Isi referensi tidak cocok dengan JIG {self.active_jig}. Kalibrasi ulang dengan tombol K.")
                return

            self.reference = data
            print(f"[OK] Referensi JIG {self.active_jig} dimuat: {self.reference_file}")
        except Exception as exc:
            print(f"[PERINGATAN] Referensi JIG {self.active_jig} gagal dibaca: {exc}")
            self.reference = None

    def has_reference(self, jig_number):
        """Mengecek apakah file referensi untuk jig tertentu tersedia."""
        return os.path.exists(self.reference_files[int(jig_number)])

    def switch_jig(self, jig_number):
        """Mengganti profil jig tanpa memakai koordinat jig sebelumnya."""
        jig_number = int(jig_number)
        if jig_number not in self.reference_files:
            print(f"[PERINGATAN] JIG {jig_number} tidak tersedia.")
            return False

        self.active_jig = jig_number
        self.reference_file = self.reference_files[jig_number]
        self.load_reference()
        self.reset_tracking()
        self.last_capture_message = f"JIG {jig_number} AKTIF"
        self.last_capture_message_until = time.time() + 2.0
        status = "SIAP" if self.reference else "BELUM DIKALIBRASI"
        print(f"[JIG] JIG {jig_number} aktif | Referensi: {status}")
        return True

    def save_reference(self, detections, frame_shape):
        height, width = frame_shape[:2]
        by_component = {det["component"]: det for det in detections if not det.get("held", False)}

        missing = [component for component in self.expected_components if component not in by_component]
        if missing:
            print("Referensi tidak disimpan. Komponen belum lengkap:", ", ".join(missing))
            return False

        components = {}
        for component in self.expected_components:
            det = by_component[component]
            x, y, box_w, box_h = det["bbox"]
            center_x, center_y = det["center"]
            components[component] = {
                "center_norm": [center_x / width, center_y / height],
                "bbox_norm": [box_w / width, box_h / height],
                "angle": float(det["angle"]),
                "source_template": det["name"],
            }

        data = {
            "version": self.REFERENCE_VERSION,
            "jig_id": self.active_jig,
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "frame_size": [width, height],
            "components": components,
        }

        with open(self.reference_file, "w", encoding="utf-8") as file:
            json.dump(data, file, indent=2, ensure_ascii=False)

        self.reference = data
        self.last_capture_message = f"REFERENSI JIG {self.active_jig} TERSIMPAN"
        self.last_capture_message_until = time.time() + 2.5
        print(f"[OK] Referensi JIG {self.active_jig} tersimpan: {self.reference_file}")
        return True

    def delete_reference(self):
        """Menghapus referensi hanya untuk jig yang sedang aktif."""
        self.reference = None
        if os.path.exists(self.reference_file):
            os.remove(self.reference_file)
        self.reset_tracking()
        print(f"[INFO] Referensi JIG {self.active_jig} dihapus. Pasang PCB GOOD lengkap lalu tekan K.")

    # --------------------------------------------------------
    # DETEKSI DAN STABILISASI
    # --------------------------------------------------------
    def _reference_search_region(self, component, frame_shape):
        """Membatasi pencarian di sekitar posisi acuan agar identitas tidak tertukar."""
        if not self.reference:
            return 0, 0, frame_shape[1], frame_shape[0]

        ref = self.reference.get("components", {}).get(component)
        if not ref:
            return 0, 0, frame_shape[1], frame_shape[0]

        height, width = frame_shape[:2]
        center_x = ref["center_norm"][0] * width
        center_y = ref["center_norm"][1] * height
        ref_w = max(1.0, ref["bbox_norm"][0] * width)
        ref_h = max(1.0, ref["bbox_norm"][1] * height)

        radius_x = max(70.0, ref_w * 4.0)
        radius_y = max(70.0, ref_h * 4.0)

        x1 = max(0, int(center_x - radius_x))
        y1 = max(0, int(center_y - radius_y))
        x2 = min(width, int(center_x + radius_x))
        y2 = min(height, int(center_y + radius_y))
        return x1, y1, x2, y2

    def calculate_angle(self, roi, component):
        if roi is None or roi.size == 0:
            return self.angle_smoother.get(component, 0.0)

        work = cv2.GaussianBlur(roi, (5, 5), 0)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(4, 4))
        work = clahe.apply(work)
        edges = cv2.Canny(work, 45, 135)

        # Abaikan tepi ROI karena sering terbaca sebagai orientasi komponen.
        border = max(2, min(edges.shape[:2]) // 25)
        edges[:border, :] = 0
        edges[-border:, :] = 0
        edges[:, :border] = 0
        edges[:, -border:] = 0
        edges = cv2.morphologyEx(edges, cv2.MORPH_CLOSE, np.ones((3, 3), np.uint8), iterations=1)

        candidates = []

        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        valid_contours = [contour for contour in contours if cv2.contourArea(contour) >= 15]
        if valid_contours:
            contour = max(valid_contours, key=cv2.contourArea)
            (_, _), (rect_w, rect_h), rect_angle = cv2.minAreaRect(contour)
            if rect_w > 1 and rect_h > 1:
                if rect_w < rect_h:
                    rect_angle += 90.0
                candidates.append(normalize_axial_angle(rect_angle))

        lines = cv2.HoughLinesP(
            edges,
            1,
            np.pi / 180.0,
            threshold=max(15, min(40, min(roi.shape[:2]) // 2)),
            minLineLength=max(8, min(roi.shape[:2]) // 4),
            maxLineGap=5,
        )
        if lines is not None:
            line_angles = []
            line_weights = []

            # Bentuk keluaran HoughLinesP dapat berbeda antar versi OpenCV:
            # (N, 1, 4), (N, 4), atau kadang array datar. Ubah semuanya
            # menjadi baris [x1, y1, x2, y2] agar tidak terjadi TypeError.
            line_array = np.asarray(lines)
            if line_array.size >= 4 and line_array.size % 4 == 0:
                line_array = line_array.reshape(-1, 4)
            else:
                line_array = np.empty((0, 4), dtype=np.int32)

            for coordinates in line_array[:20]:
                x1, y1, x2, y2 = [int(value) for value in coordinates]
                dx = x2 - x1
                dy = y2 - y1
                length = math.hypot(dx, dy)
                if length < 5:
                    continue
                line_angles.append(normalize_axial_angle(math.degrees(math.atan2(dy, dx))))
                line_weights.append(length)

            if line_angles:
                expanded = []
                max_weight = max(line_weights)
                for angle, weight in zip(line_angles, line_weights):
                    repeats = max(1, int(round(5.0 * weight / max_weight)))
                    expanded.extend([angle] * repeats)
                candidates.append(axial_mean(expanded))

        y_points, x_points = np.where(edges > 0)
        if len(x_points) >= 20:
            points = np.column_stack((x_points, y_points)).astype(np.float32)
            _, eigenvectors = cv2.PCACompute(points, mean=None)
            if eigenvectors is not None:
                eigenvector_values = np.asarray(eigenvectors, dtype=np.float64).reshape(-1)
                if eigenvector_values.size >= 2:
                    vector_x = float(eigenvector_values[0])
                    vector_y = float(eigenvector_values[1])
                    candidates.append(
                        normalize_axial_angle(math.degrees(math.atan2(vector_y, vector_x)))
                    )

        if not candidates:
            return self.angle_smoother.get(component, 0.0)

        # Bila referensi sudah ada, pilih kandidat yang paling dekat dengan arah referensi.
        ref_angle = None
        if self.reference:
            ref_data = self.reference.get("components", {}).get(component)
            if ref_data:
                ref_angle = float(ref_data.get("angle", 0.0))

        if ref_angle is not None:
            candidates.sort(key=lambda value: abs(axial_angle_difference(value, ref_angle)))
            selected = axial_mean(candidates[:2])
        else:
            selected = axial_mean(candidates)

        return self.angle_smoother.update(component, selected)

    def _match_single_template(self, gray_blurred, template_meta):
        component = template_meta["component"]
        template = template_meta["image"]
        x1, y1, x2, y2 = self._reference_search_region(component, gray_blurred.shape)
        search_image = gray_blurred[y1:y2, x1:x2]

        best_score = -1.0
        best_location = None
        best_template = None
        best_scale = None

        template_h, template_w = template.shape
        for scale in TEMPLATE_SCALES:
            new_w = max(2, int(template_w * scale))
            new_h = max(2, int(template_h * scale))
            if new_w > search_image.shape[1] or new_h > search_image.shape[0]:
                continue

            interpolation = cv2.INTER_AREA if scale < 1.0 else cv2.INTER_CUBIC
            scaled = cv2.resize(template, (new_w, new_h), interpolation=interpolation)
            result = cv2.matchTemplate(search_image, scaled, cv2.TM_CCOEFF_NORMED)
            _, max_value, _, max_location = cv2.minMaxLoc(result)

            if not np.isfinite(max_value):
                continue
            if max_value > best_score:
                best_score = float(max_value)
                best_location = (max_location[0] + x1, max_location[1] + y1)
                best_template = scaled
                best_scale = scale

        if best_location is None or best_score < self.matching_threshold:
            return None

        height, width = best_template.shape
        x, y = best_location
        return {
            "name": template_meta["name"],
            "component": component,
            "condition": template_meta["condition"],
            "bbox": (int(x), int(y), int(width), int(height)),
            "center": (float(x + width / 2.0), float(y + height / 2.0)),
            "confidence": best_score,
            "scale": best_scale,
        }

    def _remove_duplicate_regions(self, detections):
        """Mencegah beberapa nama komponen memakai area fisik yang sama."""
        ordered = sorted(detections, key=lambda item: item["confidence"], reverse=True)
        filtered = []
        for candidate in ordered:
            duplicate = False
            for existing in filtered:
                if detection_iou(candidate, existing) > 0.48:
                    duplicate = True
                    break
            if not duplicate:
                filtered.append(candidate)
        return filtered

    def _best_detection_per_component(self, detections):
        best = {}
        for det in detections:
            component = det["component"]
            if component not in best or det["confidence"] > best[component]["confidence"]:
                best[component] = det
        return list(best.values())

    def _update_tracks(self, raw_detections, gray):
        detected_by_component = {det["component"]: det for det in raw_detections}
        stable_detections = []

        for component in self.expected_components:
            track = self.tracks.setdefault(
                component,
                {"seen": 0, "missed": HOLD_MISSING_FRAMES + 1, "last": None},
            )
            current = detected_by_component.get(component)

            if current is not None:
                track["seen"] = min(track["seen"] + 1, 1000)
                track["missed"] = 0

                x, y, width, height = current["bbox"]
                roi = gray[y : y + height, x : x + width]
                angle = self.calculate_angle(roi, component)
                center = self.position_smoother.update(component, current["center"])

                smooth_x = int(round(center[0] - width / 2.0))
                smooth_y = int(round(center[1] - height / 2.0))
                smooth_x = max(0, min(gray.shape[1] - width, smooth_x))
                smooth_y = max(0, min(gray.shape[0] - height, smooth_y))

                current = dict(current)
                current["center"] = center
                current["bbox"] = (smooth_x, smooth_y, width, height)
                current["angle"] = float(angle)
                current["held"] = False
                current["stable"] = track["seen"] >= MIN_DETECTION_FRAMES
                current["seen_frames"] = track["seen"]
                current["position_std"] = self.position_smoother.get_std(component)
                current["angle_std"] = self.angle_smoother.get_std(component)
                track["last"] = current

                if current["stable"]:
                    stable_detections.append(current)
            else:
                track["seen"] = max(0, track["seen"] - 1)
                track["missed"] += 1

                if track["last"] is not None and track["missed"] <= HOLD_MISSING_FRAMES:
                    held = dict(track["last"])
                    held["held"] = True
                    held["stable"] = True
                    held["missed_frames"] = track["missed"]
                    stable_detections.append(held)
                elif track["missed"] > HOLD_MISSING_FRAMES:
                    track["last"] = None
                    self.position_smoother.clear(component)
                    self.angle_smoother.clear(component)

        return stable_detections

    def detect_components(self, frame):
        if not self.templates:
            return []

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        gray_blurred = cv2.GaussianBlur(gray, (5, 5), 0)

        matches = []
        for template_meta in self.templates:
            match = self._match_single_template(gray_blurred, template_meta)
            if match is not None:
                matches.append(match)

        matches = self._best_detection_per_component(matches)
        matches = self._remove_duplicate_regions(matches)
        return self._update_tracks(matches, gray)

    # --------------------------------------------------------
    # ANALISIS INSPEKSI
    # --------------------------------------------------------
    def analyze_board(self, detections, frame_shape):
        height, width = frame_shape[:2]
        by_component = {det["component"]: det for det in detections}
        present = sorted(by_component)
        missing = [component for component in self.expected_components if component not in by_component]

        expected_count = len(self.expected_components)
        presence_ratio = len(present) / expected_count if expected_count else 0.0
        complete = expected_count > 0 and not missing
        complete_live = complete and all(not by_component[name].get("held", False) for name in present)

        # Minimal satu komponen yang terdeteksi sudah dianggap sebagai PCB
        # yang sedang berada di jig. Ini memungkinkan auto-scan berjalan pada
        # PCB dengan banyak komponen hilang.
        board_detected = len(present) >= AUTO_CAPTURE_MIN_DETECTED_COMPONENTS
        if board_detected:
            self.board_present_counter += 1
        else:
            self.board_present_counter = 0

        component_results = []
        any_fault = False
        # all_stable dipakai untuk PCB GOOD yang wajib lengkap.
        all_stable = complete_live
        # detected_stable tetap dapat True pada PCB NOT GOOD dengan komponen hilang,
        # selama seluruh komponen yang masih terdeteksi benar-benar stabil.
        detected_stable = len(present) > 0

        for component in self.expected_components:
            det = by_component.get(component)
            if det is None:
                continue

            result = dict(det)
            result.update(
                {
                    "tilt_deg": None,
                    "setpoint_angle_deg": None,
                    "actual_angle_deg": float(det["angle"]),
                    "dx_px": None,
                    "dy_px": None,
                    "shift_px": None,
                    "shift_percent": None,
                    "shift_tolerance_px": None,
                    "angle_fault": False,
                    "angle_ignored": False,
                    "angle_info_only": is_round_component(component),
                    "shift_fault": False,
                    "template_fault": det["condition"] == "NOT_GOOD",
                }
            )

            angle_reading_stable = (
                True
                if is_round_component(component)
                else det.get("angle_std", 999.0) <= MAX_ANGLE_STD_DEG
            )
            current_detection_stable = (
                not det.get("held", False)
                and det.get("position_std", 999.0) <= MAX_POSITION_STD_PX
                and angle_reading_stable
                and det.get("seen_frames", 0) >= MIN_DETECTION_FRAMES
            )
            all_stable = all_stable and current_detection_stable
            detected_stable = detected_stable and current_detection_stable

            if self.reference:
                ref = self.reference.get("components", {}).get(component)
                if ref:
                    ref_center_x = float(ref["center_norm"][0]) * width
                    ref_center_y = float(ref["center_norm"][1]) * height
                    ref_box_w = max(1.0, float(ref["bbox_norm"][0]) * width)
                    ref_box_h = max(1.0, float(ref["bbox_norm"][1]) * height)
                    ref_diagonal = math.hypot(ref_box_w, ref_box_h)

                    center_x, center_y = det["center"]
                    dx = float(center_x - ref_center_x)
                    dy = float(center_y - ref_center_y)
                    shift = math.hypot(dx, dy)
                    shift_percent = 100.0 * shift / max(ref_diagonal, 1.0)
                    shift_tolerance = max(
                        MIN_SHIFT_TOLERANCE_PX,
                        ref_diagonal * SHIFT_TOLERANCE_RATIO,
                    )
                    # Selisih sudut dihitung untuk SEMUA komponen agar
                    # keterangannya selalu dapat ditampilkan.
                    tilt = axial_angle_difference(
                        det["angle"],
                        float(ref.get("angle", 0.0)),
                    )

                    # C1/C2 tetap hanya menggunakan sudut sebagai informasi
                    # karena bentuk bulat mudah berubah akibat pantulan cahaya.
                    # Komponen lain memakai sudut untuk keputusan GOOD/NOT GOOD.
                    if is_round_component(component):
                        angle_fault = False
                    else:
                        angle_fault = abs(tilt) >= ANGLE_TOLERANCE_DEG

                    result.update(
                        {
                            "tilt_deg": tilt,
                            "setpoint_angle_deg": float(ref.get("angle", 0.0)),
                            "actual_angle_deg": float(det["angle"]),
                            "dx_px": dx,
                            "dy_px": dy,
                            "shift_px": shift,
                            "shift_percent": shift_percent,
                            "shift_tolerance_px": shift_tolerance,
                            "angle_fault": angle_fault,
                            "shift_fault": shift > shift_tolerance,
                        }
                    )

            result["fault"] = bool(
                result["template_fault"] or result["angle_fault"] or result["shift_fault"]
            )
            any_fault = any_fault or result["fault"]
            component_results.append(result)

        evaluation_mature = complete or self.board_present_counter >= 12

        if expected_count == 0:
            decision = "TEMPLATE KOSONG"
        elif not board_detected:
            decision = "MENUNGGU PCB"
        elif not self.reference:
            decision = "KALIBRASI"
        elif evaluation_mature and (missing or any_fault):
            decision = "NOT GOOD"
        elif complete:
            decision = "GOOD"
        else:
            decision = "MEMBACA"

        fault_for_alarm = decision == "NOT GOOD" and evaluation_mature
        self.alarm.update_fault(fault_for_alarm)

        return {
            "decision": decision,
            "expected_count": expected_count,
            "detected_count": len(present),
            "presence_ratio": presence_ratio,
            "board_detected": board_detected,
            "present_components": present,
            "missing_components": missing,
            "complete": complete,
            "complete_live": complete_live,
            "all_stable": all_stable,
            "detected_stable": detected_stable,
            "components": component_results,
            "reference_ready": self.reference is not None,
            "evaluation_mature": evaluation_mature,
            "angle_tolerance_deg": ANGLE_TOLERANCE_DEG,
            "angle_good_rule": f"abs(kemiringan) < {ANGLE_TOLERANCE_DEG:.1f} derajat",
            "angle_not_good_rule": f"abs(kemiringan) >= {ANGLE_TOLERANCE_DEG:.1f} derajat",
            "min_shift_tolerance_px": MIN_SHIFT_TOLERANCE_PX,
        }

    # --------------------------------------------------------
    # AUTO CAPTURE
    # --------------------------------------------------------
    def _capture_signature(self, analysis, decision):
        """Identitas hasil untuk mendeteksi perubahan selama timer stabil."""
        component_status = tuple(
            sorted(
                (
                    item["component"],
                    bool(item.get("template_fault", False)),
                    bool(item.get("angle_fault", False)),
                    bool(item.get("shift_fault", False)),
                )
                for item in analysis.get("components", [])
            )
        )
        return (
            decision,
            tuple(sorted(analysis.get("missing_components", []))),
            component_status,
        )

    def update_auto_capture_state(self, analysis):
        """Scan 5 detik lalu tunggu hasil stabil 5 detik."""
        now = time.time()
        detected_count = int(analysis.get("detected_count", 0))
        board_visible = detected_count >= AUTO_CAPTURE_MIN_DETECTED_COMPONENTS
        reference_ready = bool(
            self.reference and analysis.get("reference_ready", False)
        )

        # Sesudah capture, tunggu PCB diangkat.
        if self.auto_capture_locked:
            self.auto_capture_phase = "LOCKED"
            if detected_count == 0:
                self.incomplete_counter += 1
                if self.incomplete_counter >= AUTO_CAPTURE_UNLOCK_FRAMES:
                    self.reset_auto_capture_lock()
                    print(
                        "[INFO] PCB diangkat. Sistem siap untuk PCB berikutnya."
                    )
            else:
                self.incomplete_counter = 0
            return None

        # Referensi diperlukan untuk menentukan set point dan komponen hilang.
        if not reference_ready:
            if self.auto_capture_phase != "WAITING":
                self.reset_auto_capture_lock()
            return None

        # PCB hilang saat proses berlangsung.
        if not board_visible:
            if self.auto_capture_phase in ("SCANNING", "STABILIZING"):
                self.auto_capture_empty_counter += 1
                if self.auto_capture_empty_counter >= AUTO_CAPTURE_EMPTY_CANCEL_FRAMES:
                    print("[AUTO] PCB tidak terlihat. Proses dibatalkan.")
                    self.reset_auto_capture_lock()
            return None

        self.auto_capture_empty_counter = 0

        # TAHAP 1: SCANNING 5 DETIK
        if self.auto_capture_phase == "WAITING":
            self.auto_capture_phase = "SCANNING"
            self.auto_scan_started_at = now
            self.auto_scan_remaining_seconds = INITIAL_SCAN_SECONDS
            self.auto_capture_candidate_decision = "MEMBACA"
            print(
                f"[AUTO] PCB terdeteksi. Scanning "
                f"{INITIAL_SCAN_SECONDS:.0f} detik dimulai."
            )
            return None

        if self.auto_capture_phase == "SCANNING":
            if self.auto_scan_started_at is None:
                self.auto_scan_started_at = now

            elapsed_scan = max(0.0, now - self.auto_scan_started_at)
            self.auto_scan_remaining_seconds = max(
                0.0,
                INITIAL_SCAN_SECONDS - elapsed_scan,
            )

            if elapsed_scan < INITIAL_SCAN_SECONDS:
                return None

            self.auto_capture_phase = "STABILIZING"
            self.auto_stable_started_at = None
            self.auto_capture_candidate_signature = None
            self.auto_capture_remaining_seconds = STABLE_CAPTURE_SECONDS
            print(
                "[AUTO] Scanning selesai. Menunggu pembacaan stabil."
            )

        # Tentukan kandidat hasil berdasarkan pembacaan terbaru.
        component_fault = any(
            bool(item.get("fault", False))
            for item in analysis.get("components", [])
        )
        has_missing = bool(analysis.get("missing_components", []))
        complete = bool(analysis.get("complete", False))

        if complete and not has_missing and not component_fault:
            candidate_decision = "GOOD"
            stable_ready = bool(
                analysis.get("complete_live", False)
                and analysis.get("all_stable", False)
            )
        else:
            candidate_decision = "NOT GOOD"
            stable_ready = bool(
                detected_count >= AUTO_CAPTURE_MIN_DETECTED_COMPONENTS
                and analysis.get("detected_stable", False)
            )

        self.auto_capture_candidate_decision = candidate_decision

        # Pembacaan posisi/sudut belum stabil.
        if not stable_ready:
            self.auto_stable_started_at = None
            self.auto_capture_candidate_signature = None
            self.auto_capture_remaining_seconds = STABLE_CAPTURE_SECONDS
            return None

        signature = self._capture_signature(analysis, candidate_decision)

        # Jika daftar komponen hilang atau fault berubah, timer kembali ke 5 detik.
        if signature != self.auto_capture_candidate_signature:
            self.auto_capture_candidate_signature = signature
            self.auto_stable_started_at = now
            self.auto_capture_remaining_seconds = STABLE_CAPTURE_SECONDS
            print(
                f"[AUTO] Hasil {candidate_decision} terbaca. "
                f"Tunggu stabil {STABLE_CAPTURE_SECONDS:.0f} detik."
            )
            return None

        if self.auto_stable_started_at is None:
            self.auto_stable_started_at = now

        elapsed_stable = max(0.0, now - self.auto_stable_started_at)
        self.auto_capture_remaining_seconds = max(
            0.0,
            STABLE_CAPTURE_SECONDS - elapsed_stable,
        )
        self.auto_capture_counter = int(round(elapsed_stable * 10.0))

        if elapsed_stable < STABLE_CAPTURE_SECONDS:
            return None

        # Capture hasil yang sudah stabil.
        analysis["decision"] = candidate_decision
        analysis["evaluation_mature"] = True
        self.alarm.update_fault(candidate_decision == "NOT GOOD")

        self.auto_capture_locked = True
        self.auto_capture_phase = "LOCKED"
        self.incomplete_counter = 0
        self.auto_capture_remaining_seconds = 0.0

        print(
            f"[AUTO] Capture otomatis {candidate_decision} dijalankan."
        )
        return "CAPTURE"

    def reset_auto_capture_lock(self):
        self.auto_capture_counter = 0
        self.auto_capture_locked = False
        self.incomplete_counter = 0
        self.auto_capture_phase = "WAITING"
        self.auto_scan_started_at = None
        self.auto_stable_started_at = None
        self.auto_capture_candidate_signature = None
        self.auto_capture_candidate_decision = None
        self.auto_scan_remaining_seconds = INITIAL_SCAN_SECONDS
        self.auto_capture_remaining_seconds = STABLE_CAPTURE_SECONDS
        self.auto_capture_empty_counter = 0

    def unlock_auto_capture(self):
        self.reset_auto_capture_lock()
        self.last_capture_message = "AUTO CAPTURE DIBUKA"
        self.last_capture_message_until = time.time() + 1.5
        print("[INFO] Auto capture dibuka secara manual.")

    def _safe_crop(self, frame, bbox, padding=DATASET_CROP_PADDING):
        """Mengambil crop dengan padding tanpa keluar dari batas gambar."""
        if frame is None or frame.size == 0:
            return None
        x, y, width, height = [int(round(value)) for value in bbox]
        x1 = max(0, x - padding)
        y1 = max(0, y - padding)
        x2 = min(frame.shape[1], x + width + padding)
        y2 = min(frame.shape[0], y + height + padding)
        if x2 <= x1 or y2 <= y1:
            return None
        crop = frame[y1:y2, x1:x2]
        return crop.copy() if crop.size > 0 else None

    def _missing_component_bbox(self, component, frame_shape):
        """ROI acuan komponen hilang agar dapat disimpan sebagai sampel MISSING."""
        if not self.reference:
            return None
        ref = self.reference.get("components", {}).get(component)
        if not ref:
            return None
        height, width = frame_shape[:2]
        center_x = float(ref["center_norm"][0]) * width
        center_y = float(ref["center_norm"][1]) * height
        box_w = max(4, int(round(float(ref["bbox_norm"][0]) * width)))
        box_h = max(4, int(round(float(ref["bbox_norm"][1]) * height)))
        return (
            int(round(center_x - box_w / 2.0)),
            int(round(center_y - box_h / 2.0)),
            box_w,
            box_h,
        )

    def save_learning_dataset(self, raw_frame, annotated_frame, analysis, timestamp, trigger):
        """
        Menyimpan dataset otomatis:
        - gambar PCB mentah menurut GOOD/NOT_GOOD;
        - gambar anotasi untuk pemeriksaan;
        - crop tiap komponen GOOD/NOT_GOOD;
        - ROI komponen hilang ke kelas MISSING;
        - JSON dan CSV metadata.
        """
        board_status = "GOOD" if analysis["decision"] == "GOOD" else "NOT_GOOD"
        os.makedirs(self.dataset_folder, exist_ok=True)

        board_dir = os.path.join(self.dataset_folder, "PCB", board_status)
        annotated_dir = os.path.join(self.dataset_folder, "PCB_ANOTASI", board_status)
        metadata_dir = os.path.join(self.dataset_folder, "ANOTASI_JSON", board_status)
        for folder in (board_dir, annotated_dir, metadata_dir):
            os.makedirs(folder, exist_ok=True)

        base_name = f"{timestamp}_JIG_{self.active_jig}_{safe_name(trigger)}_{board_status}"
        board_filename = base_name + ".jpg"
        board_path = os.path.join(board_dir, board_filename)
        annotated_path = os.path.join(annotated_dir, base_name + "_anotasi.jpg")
        annotation_path = os.path.join(metadata_dir, base_name + ".json")

        cv2.imwrite(board_path, raw_frame, [cv2.IMWRITE_JPEG_QUALITY, JPEG_QUALITY])
        cv2.imwrite(annotated_path, annotated_frame, [cv2.IMWRITE_JPEG_QUALITY, JPEG_QUALITY])

        objects = []
        crop_paths = []
        for item in analysis.get("components", []):
            component = safe_name(item["component"])
            component_status = "NOT_GOOD" if item.get("fault", False) else "GOOD"
            crop = self._safe_crop(raw_frame, item["bbox"])
            crop_relative = None
            if crop is not None:
                crop_dir = os.path.join(
                    self.dataset_folder,
                    "KOMPONEN",
                    component,
                    component_status,
                )
                os.makedirs(crop_dir, exist_ok=True)
                crop_name = f"{base_name}_{component}_{component_status}.jpg"
                crop_path = os.path.join(crop_dir, crop_name)
                cv2.imwrite(crop_path, crop, [cv2.IMWRITE_JPEG_QUALITY, JPEG_QUALITY])
                crop_relative = os.path.relpath(crop_path, self.dataset_folder)
                crop_paths.append(crop_path)

            x, y, box_w, box_h = [int(round(value)) for value in item["bbox"]]
            objects.append(
                {
                    "component": item["component"],
                    "status": component_status,
                    "bbox_xywh": [x, y, box_w, box_h],
                    "confidence": round(float(item["confidence"]), 4),
                    "template": item["name"],
                    "crop_file": crop_relative,
                }
            )

        # Untuk komponen hilang, simpan area tempat komponen seharusnya berada.
        for component_name in analysis.get("missing_components", []):
            missing_bbox = self._missing_component_bbox(component_name, raw_frame.shape)
            crop_relative = None
            if missing_bbox is not None:
                missing_crop = self._safe_crop(raw_frame, missing_bbox)
                if missing_crop is not None:
                    component = safe_name(component_name)
                    crop_dir = os.path.join(
                        self.dataset_folder,
                        "KOMPONEN",
                        component,
                        "MISSING",
                    )
                    os.makedirs(crop_dir, exist_ok=True)
                    crop_name = f"{base_name}_{component}_MISSING.jpg"
                    crop_path = os.path.join(crop_dir, crop_name)
                    cv2.imwrite(crop_path, missing_crop, [cv2.IMWRITE_JPEG_QUALITY, JPEG_QUALITY])
                    crop_relative = os.path.relpath(crop_path, self.dataset_folder)
                    crop_paths.append(crop_path)

            objects.append(
                {
                    "component": component_name,
                    "status": "MISSING",
                    "bbox_xywh": None if missing_bbox is None else list(missing_bbox),
                    "confidence": 0.0,
                    "template": None,
                    "crop_file": crop_relative,
                }
            )

        annotation = {
            "timestamp": datetime.now().isoformat(timespec="milliseconds"),
            "image_file": os.path.relpath(board_path, self.dataset_folder),
            "annotated_file": os.path.relpath(annotated_path, self.dataset_folder),
            "image_width": int(raw_frame.shape[1]),
            "image_height": int(raw_frame.shape[0]),
            "jig_id": self.active_jig,
            "board_status": board_status,
            "decision": analysis["decision"],
            "missing_components": list(analysis.get("missing_components", [])),
            "objects": objects,
            "catatan": "Label otomatis wajib diperiksa sebelum dipakai melatih model.",
        }
        with open(annotation_path, "w", encoding="utf-8") as file:
            json.dump(annotation, file, indent=2, ensure_ascii=False)

        csv_path = os.path.join(self.dataset_folder, "metadata_dataset.csv")
        csv_exists = os.path.exists(csv_path) and os.path.getsize(csv_path) > 0
        with open(csv_path, "a", newline="", encoding="utf-8") as file:
            fieldnames = [
                "timestamp",
                "image_file",
                "jig_id",
                "board_status",
                "decision",
                "detected_count",
                "expected_count",
                "missing_components",
                "trigger",
            ]
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            if not csv_exists:
                writer.writeheader()
            writer.writerow(
                {
                    "timestamp": annotation["timestamp"],
                    "image_file": annotation["image_file"],
                    "jig_id": self.active_jig,
                    "board_status": board_status,
                    "decision": analysis["decision"],
                    "detected_count": analysis["detected_count"],
                    "expected_count": analysis["expected_count"],
                    "missing_components": "|".join(analysis.get("missing_components", [])),
                    "trigger": trigger,
                }
            )

        # Daftar kelas membantu saat dataset akan dilabel ulang untuk YOLO.
        classes_path = os.path.join(self.dataset_folder, "daftar_komponen.txt")
        with open(classes_path, "w", encoding="utf-8") as file:
            for component in self.expected_components:
                file.write(component + "\n")

        print(f"[DATASET PCB] {board_path}")
        print(f"[DATASET JSON] {annotation_path}")
        print(f"[DATASET CROP] {len(crop_paths)} gambar komponen tersimpan")
        return {
            "board": board_path,
            "annotated": annotated_path,
            "annotation": annotation_path,
            "metadata_csv": csv_path,
            "crops": crop_paths,
        }


    # --------------------------------------------------------
    # PENYIMPANAN EXCEL DATA KOMPONEN LENGKAP PER JIG
    # --------------------------------------------------------
    def _excel_file_path(self):
        """File Excel dibuat terpisah untuk Jig 1 dan Jig 2."""
        jig_folder = os.path.join(
            self.excel_base_folder,
            f"JIG_{self.active_jig}",
        )
        os.makedirs(jig_folder, exist_ok=True)

        return os.path.join(
            jig_folder,
            f"Data_Komponen_Lengkap_JIG_{self.active_jig}.xlsx",
        )

    @staticmethod
    def _excel_headers():
        """Judul kolom yang dipakai oleh seluruh file Excel."""
        summary_headers = [
            "ID INSPEKSI",
            "TANGGAL",
            "WAKTU",
            "JIG",
            "PEMICU",
            "STATUS PCB",
            "JUMLAH TERDETEKSI",
            "TOTAL KOMPONEN",
            "JUMLAH GOOD",
            "JUMLAH NOT GOOD",
            "JUMLAH HILANG",
            "DAFTAR KOMPONEN GOOD",
            "DAFTAR KOMPONEN NOT GOOD",
            "KEMIRINGAN SEMUA KOMPONEN (°)",
            "PERGESERAN SEMUA KOMPONEN (px)",
            "DAFTAR KOMPONEN HILANG",
            "GAMBAR RAW",
            "GAMBAR HASIL",
            "LAPORAN JSON",
            "LAPORAN TXT",
        ]

        component_headers = [
            "ID INSPEKSI",
            "TANGGAL",
            "WAKTU",
            "JIG",
            "STATUS PCB",
            "PEMICU",
            "NAMA TEMPLATE TERDETEKSI",
            "ID KOMPONEN",
            "KONDISI TEMPLATE",
            "FORMAT DATA KOMPONEN",
            "TERDETEKSI",
            "STATUS AKHIR KOMPONEN",
            "KETERANGAN",
            "CONFIDENCE (%)",
            "SET POINT SUDUT (°)",
            "SUDUT AKTUAL (°)",
            "KEMIRINGAN DARI SET POINT (°)",
            "STATUS KEMIRINGAN",
            "SUDUT HANYA INFORMASI",
            "DX (px)",
            "DY (px)",
            "PERGESERAN (px)",
            "BATAS PERGESERAN (px)",
            "STATUS PERGESERAN",
            "KOMPONEN HILANG",
            "GAMBAR RAW",
            "GAMBAR HASIL",
        ]

        template_headers = [
            "NO",
            "NAMA TEMPLATE",
            "FORMAT MAPPING",
            "ID KOMPONEN",
            "KONDISI TEMPLATE",
            "LOKASI FILE TEMPLATE",
        ]

        return summary_headers, component_headers, template_headers

    @staticmethod
    def _style_excel_header(sheet, row_number, max_column):
        """Format judul sheet."""
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        border_side = Side(style="thin", color="B7C9D6")

        for column in range(1, max_column + 1):
            cell = sheet.cell(row=row_number, column=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = Border(
                left=border_side,
                right=border_side,
                top=border_side,
                bottom=border_side,
            )

        sheet.row_dimensions[row_number].height = 34

    @staticmethod
    def _style_excel_data_row(sheet, row_number, status_column):
        """Warna status pada baris data."""
        border_side = Side(style="thin", color="D9E2F3")
        status = str(
            sheet.cell(
                row=row_number,
                column=status_column,
            ).value or ""
        ).upper()

        if status == "GOOD":
            status_fill = PatternFill("solid", fgColor="E2F0D9")
            status_font = Font(color="006100", bold=True)
        elif status in ("NOT GOOD", "HILANG"):
            status_fill = PatternFill("solid", fgColor="FCE4D6")
            status_font = Font(color="9C0006", bold=True)
        else:
            status_fill = PatternFill("solid", fgColor="FFF2CC")
            status_font = Font(color="7F6000", bold=True)

        for cell in sheet[row_number]:
            cell.border = Border(
                left=border_side,
                right=border_side,
                top=border_side,
                bottom=border_side,
            )
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        status_cell = sheet.cell(
            row=row_number,
            column=status_column,
        )
        status_cell.fill = status_fill
        status_cell.font = status_font

    @staticmethod
    def _set_excel_column_widths(sheet, widths):
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[
                get_column_letter(index)
            ].width = width

    def _refresh_template_sheet(self, workbook):
        """Menulis daftar template persis seperti yang dibaca program."""
        if "Daftar Template" in workbook.sheetnames:
            sheet = workbook["Daftar Template"]
            sheet.delete_rows(1, sheet.max_row)
        else:
            sheet = workbook.create_sheet("Daftar Template")

        _, _, template_headers = self._excel_headers()
        sheet.append(template_headers)
        self._style_excel_header(
            sheet,
            1,
            len(template_headers),
        )

        for number, item in enumerate(self.templates, start=1):
            condition_text = (
                "NOT GOOD"
                if item["condition"] == "NOT_GOOD"
                else "GOOD"
            )
            mapping_text = (
                f"{item['name']} -> "
                f"{item['component']} ({condition_text})"
            )

            sheet.append(
                [
                    number,
                    item["name"],
                    mapping_text,
                    item["component"],
                    condition_text,
                    os.path.abspath(item["path"]),
                ]
            )

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(template_headers))}"
            f"{max(1, sheet.max_row)}"
        )
        self._set_excel_column_widths(
            sheet,
            [8, 28, 48, 22, 20, 58],
        )

        for row_number in range(2, sheet.max_row + 1):
            for cell in sheet[row_number]:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
                cell.border = Border(
                    left=Side(style="thin", color="D9E2F3"),
                    right=Side(style="thin", color="D9E2F3"),
                    top=Side(style="thin", color="D9E2F3"),
                    bottom=Side(style="thin", color="D9E2F3"),
                )

    @staticmethod
    def _component_sheet_name(component):
        """Nama sheet aman, maksimal 31 karakter sesuai aturan Excel."""
        cleaned = re.sub(
            r"[:\\/?*\[\]]+",
            "_",
            str(component).strip().upper(),
        )
        cleaned = cleaned[:31]
        return cleaned or "KOMPONEN"

    def _ensure_component_sheet(self, workbook, component):
        """Memastikan sheet khusus komponen tersedia dalam workbook yang sama."""
        sheet_name = self._component_sheet_name(component)
        _, component_headers, _ = self._excel_headers()

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(component_headers)
            self._style_excel_header(
                sheet,
                1,
                len(component_headers),
            )
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = (
                f"A1:{get_column_letter(len(component_headers))}1"
            )
            self._set_excel_column_widths(
                sheet,
                [
                    26, 13, 12, 9, 16, 12, 30, 22, 21,
                    48, 13, 24, 58, 17, 22, 20, 30, 22,
                    24, 13, 13, 19, 24, 22, 18, 48, 48,
                ],
            )

        return sheet

    def _ensure_all_component_sheets(self, workbook):
        """Membuat satu sheet untuk setiap komponen yang diharapkan."""
        result = {}
        for component in self.expected_components:
            result[component] = self._ensure_component_sheet(
                workbook,
                component,
            )
        return result

    def _append_component_row_to_sheet(
        self,
        sheet,
        component_row,
    ):
        """Menambahkan baris lengkap ke sheet komponen tertentu."""
        sheet.append(component_row)
        row_number = sheet.max_row

        sheet.cell(
            row=row_number,
            column=2,
        ).number_format = "dd/mm/yyyy"
        sheet.cell(
            row=row_number,
            column=3,
        ).number_format = "hh:mm:ss"

        self._format_excel_measurement_cells(
            sheet,
            row_number,
        )
        self._style_excel_data_row(
            sheet,
            row_number,
            12,
        )

        for column in (26, 27):
            link_cell = sheet.cell(
                row=row_number,
                column=column,
            )
            if link_cell.value:
                link_cell.hyperlink = str(link_cell.value)
                link_cell.style = "Hyperlink"

        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(sheet.max_column)}"
            f"{sheet.max_row}"
        )

    def _migrate_existing_rows_to_component_sheets(
        self,
        workbook,
    ):
        """Memindahkan data lama dari Data Komponen ke sheet per komponen.

        Data lama tidak dihapus. Fungsi hanya menyalin baris yang belum
        terdapat pada sheet komponen berdasarkan ID inspeksi.
        """
        if "Data Komponen" not in workbook.sheetnames:
            return

        combined_sheet = workbook["Data Komponen"]
        component_sheets = self._ensure_all_component_sheets(
            workbook
        )

        existing_ids = {}
        for component, sheet in component_sheets.items():
            ids = set()
            for row_number in range(2, sheet.max_row + 1):
                inspection_id = sheet.cell(
                    row=row_number,
                    column=1,
                ).value
                if inspection_id:
                    ids.add(str(inspection_id))
            existing_ids[component] = ids

        for row_number in range(2, combined_sheet.max_row + 1):
            inspection_id = combined_sheet.cell(
                row=row_number,
                column=1,
            ).value
            component = combined_sheet.cell(
                row=row_number,
                column=8,
            ).value

            if not inspection_id or not component:
                continue

            component = str(component)
            if component not in component_sheets:
                component_sheets[component] = (
                    self._ensure_component_sheet(
                        workbook,
                        component,
                    )
                )
                existing_ids[component] = set()

            if str(inspection_id) in existing_ids[component]:
                continue

            component_row = [
                combined_sheet.cell(
                    row=row_number,
                    column=column,
                ).value
                for column in range(
                    1,
                    combined_sheet.max_column + 1,
                )
            ]

            self._append_component_row_to_sheet(
                component_sheets[component],
                component_row,
            )
            existing_ids[component].add(str(inspection_id))

    def _create_excel_workbook(self, excel_path):
        """Membuat workbook untuk data komponen Jig aktif."""
        workbook = ExcelWorkbook()

        summary_sheet = workbook.active
        summary_sheet.title = "Ringkasan PCB"
        component_sheet = workbook.create_sheet("Data Komponen")
        workbook.create_sheet("Daftar Template")
        statistic_sheet = workbook.create_sheet("Statistik")

        (
            summary_headers,
            component_headers,
            _,
        ) = self._excel_headers()

        summary_sheet.append(summary_headers)
        component_sheet.append(component_headers)

        self._style_excel_header(
            summary_sheet,
            1,
            len(summary_headers),
        )
        self._style_excel_header(
            component_sheet,
            1,
            len(component_headers),
        )

        summary_sheet.freeze_panes = "A2"
        component_sheet.freeze_panes = "A2"

        summary_sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(summary_headers))}1"
        )
        component_sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(component_headers))}1"
        )

        self._set_excel_column_widths(
            summary_sheet,
            [
                26, 13, 12, 9, 12, 16, 18, 17, 14, 18,
                15, 34, 38, 38, 38, 38, 48, 48, 48, 48,
            ],
        )

        self._set_excel_column_widths(
            component_sheet,
            [
                26, 13, 12, 9, 16, 12, 30, 22, 21, 48,
                13, 24, 48, 17, 22, 20, 30, 22, 24, 13,
                13, 19, 24, 22, 18, 48, 48,
            ],
        )

        # Tetap satu workbook, tetapi setiap komponen mempunyai sheet sendiri.
        self._ensure_all_component_sheets(workbook)

        # Daftar seluruh template:
        # C1 GOOD -> C1 (GOOD)
        # Connector NOT GOOD -> CONNECTOR (NOT GOOD)
        self._refresh_template_sheet(workbook)

        statistic_sheet["A1"] = (
            f"STATISTIK DATA KOMPONEN JIG {self.active_jig}"
        )
        statistic_sheet.merge_cells("A1:D1")
        statistic_sheet["A1"].fill = PatternFill(
            "solid",
            fgColor="1F4E78",
        )
        statistic_sheet["A1"].font = Font(
            color="FFFFFF",
            bold=True,
            size=15,
        )
        statistic_sheet["A1"].alignment = Alignment(
            horizontal="center",
        )

        statistic_sheet.append([])
        statistic_sheet.append(["KETERANGAN", "NILAI"])
        self._style_excel_header(statistic_sheet, 3, 2)

        statistic_rows = [
            [
                "Total inspeksi",
                "=MAX(COUNTA('Ringkasan PCB'!A:A)-1,0)",
            ],
            [
                "PCB GOOD",
                '=COUNTIF(\'Ringkasan PCB\'!F:F,"GOOD")',
            ],
            [
                "PCB NOT GOOD",
                '=COUNTIF(\'Ringkasan PCB\'!F:F,"NOT GOOD")',
            ],
            [
                "Total data komponen",
                "=MAX(COUNTA('Data Komponen'!A:A)-1,0)",
            ],
            [
                "Komponen GOOD",
                '=COUNTIF(\'Data Komponen\'!L:L,"GOOD")',
            ],
            [
                "Komponen NOT GOOD",
                '=COUNTIF(\'Data Komponen\'!L:L,"NOT GOOD")',
            ],
            [
                "Komponen HILANG",
                '=COUNTIF(\'Data Komponen\'!L:L,"HILANG")',
            ],
            [
                "Kemiringan NOT GOOD",
                '=COUNTIF(\'Data Komponen\'!R:R,"NOT GOOD")',
            ],
            [
                "Pergeseran NOT GOOD",
                '=COUNTIF(\'Data Komponen\'!X:X,"NOT GOOD")',
            ],
            [
                "Terakhir diperbarui",
                "=NOW()",
            ],
        ]

        for row in statistic_rows:
            statistic_sheet.append(row)

        statistic_sheet["B13"].number_format = (
            "dd/mm/yyyy hh:mm:ss"
        )
        statistic_sheet.column_dimensions["A"].width = 32
        statistic_sheet.column_dimensions["B"].width = 24
        statistic_sheet.freeze_panes = "A4"

        for row_number in range(4, statistic_sheet.max_row + 1):
            statistic_sheet.cell(
                row=row_number,
                column=1,
            ).font = Font(bold=True)

        workbook.save(excel_path)
        return workbook

    @staticmethod
    def _component_reason(item):
        """Menyusun keterangan lengkap dengan nilai dan satuannya."""
        reasons = []

        template_status = (
            "NOT GOOD"
            if item.get("condition") == "NOT_GOOD"
            else "GOOD"
        )
        reasons.append(f"TEMPLATE {template_status}")

        tilt = item.get("tilt_deg")
        if tilt is None:
            reasons.append("KEMIRINGAN N/A")
        else:
            if item.get("angle_info_only", False):
                angle_status = "INFORMASI"
            elif item.get("angle_fault", False):
                angle_status = "NOT GOOD"
            else:
                angle_status = "GOOD"

            reasons.append(
                f"KEMIRINGAN {float(tilt):+.2f}° "
                f"({angle_status})"
            )

        shift = item.get("shift_px")
        if shift is None:
            reasons.append("PERGESERAN N/A")
        else:
            shift_status = (
                "NOT GOOD"
                if item.get("shift_fault", False)
                else "GOOD"
            )
            reasons.append(
                f"PERGESERAN {float(shift):.2f} px "
                f"({shift_status})"
            )

        return " | ".join(reasons)

    @staticmethod
    def _format_excel_measurement_cells(component_sheet, row_number):
        """Menampilkan simbol derajat dan pixel tanpa mengubah nilai numerik."""
        degree_format = '+0.00"°";-0.00"°";0.00"°"'
        signed_pixel_format = '+0.00" px";-0.00" px";0.00" px"'
        pixel_format = '0.00" px"'

        # O, P, Q: set point, aktual, selisih kemiringan.
        for column in (15, 16, 17):
            component_sheet.cell(
                row=row_number,
                column=column,
            ).number_format = degree_format

        # T, U: DX dan DY bisa bernilai positif atau negatif.
        for column in (20, 21):
            component_sheet.cell(
                row=row_number,
                column=column,
            ).number_format = signed_pixel_format

        # V, W: besar pergeseran dan batas pergeseran selalu positif.
        for column in (22, 23):
            component_sheet.cell(
                row=row_number,
                column=column,
            ).number_format = pixel_format

    def _repair_existing_excel_measurements(self, workbook):
        """Memperbaiki unit dan keterangan pada baris Excel yang sudah ada."""
        if "Data Komponen" not in workbook.sheetnames:
            return

        component_sheet = workbook["Data Komponen"]

        # Pastikan judul memakai unit yang benar.
        component_sheet.cell(row=1, column=15).value = "SET POINT SUDUT (°)"
        component_sheet.cell(row=1, column=16).value = "SUDUT AKTUAL (°)"
        component_sheet.cell(
            row=1,
            column=17,
        ).value = "KEMIRINGAN DARI SET POINT (°)"
        component_sheet.cell(row=1, column=20).value = "DX (px)"
        component_sheet.cell(row=1, column=21).value = "DY (px)"
        component_sheet.cell(row=1, column=22).value = "PERGESERAN (px)"
        component_sheet.cell(
            row=1,
            column=23,
        ).value = "BATAS PERGESERAN (px)"

        for row_number in range(2, component_sheet.max_row + 1):
            self._format_excel_measurement_cells(
                component_sheet,
                row_number,
            )

            detected = str(
                component_sheet.cell(
                    row=row_number,
                    column=11,
                ).value or ""
            ).upper()

            if detected == "TIDAK":
                component_sheet.cell(
                    row=row_number,
                    column=13,
                ).value = (
                    "KOMPONEN HILANG / TIDAK TERDETEKSI"
                )
                continue

            template_status = str(
                component_sheet.cell(
                    row=row_number,
                    column=9,
                ).value or "GOOD"
            )
            tilt_value = component_sheet.cell(
                row=row_number,
                column=17,
            ).value
            angle_status = str(
                component_sheet.cell(
                    row=row_number,
                    column=18,
                ).value or "GOOD"
            )
            shift_value = component_sheet.cell(
                row=row_number,
                column=22,
            ).value
            shift_status = str(
                component_sheet.cell(
                    row=row_number,
                    column=24,
                ).value or "GOOD"
            )

            if isinstance(tilt_value, (int, float)):
                tilt_text = (
                    f"KEMIRINGAN {float(tilt_value):+.2f}° "
                    f"({angle_status})"
                )
            else:
                tilt_text = "KEMIRINGAN N/A"

            if isinstance(shift_value, (int, float)):
                shift_text = (
                    f"PERGESERAN {float(shift_value):.2f} px "
                    f"({shift_status})"
                )
            else:
                shift_text = "PERGESERAN N/A"

            component_sheet.cell(
                row=row_number,
                column=13,
            ).value = (
                f"TEMPLATE {template_status} | "
                f"{tilt_text} | {shift_text}"
            )

        # Perbaiki kolom ringkasan memakai data detail yang sudah ada.
        if "Ringkasan PCB" not in workbook.sheetnames:
            return

        summary_sheet = workbook["Ringkasan PCB"]
        summary_sheet.cell(
            row=1,
            column=14,
        ).value = "KEMIRINGAN SEMUA KOMPONEN (°)"
        summary_sheet.cell(
            row=1,
            column=15,
        ).value = "PERGESERAN SEMUA KOMPONEN (px)"

        detail_by_inspection = {}
        for row_number in range(2, component_sheet.max_row + 1):
            inspection_id = component_sheet.cell(
                row=row_number,
                column=1,
            ).value
            component = component_sheet.cell(
                row=row_number,
                column=8,
            ).value
            detected = str(
                component_sheet.cell(
                    row=row_number,
                    column=11,
                ).value or ""
            ).upper()

            if not inspection_id or detected == "TIDAK":
                continue

            tilt_value = component_sheet.cell(
                row=row_number,
                column=17,
            ).value
            angle_status = component_sheet.cell(
                row=row_number,
                column=18,
            ).value
            shift_value = component_sheet.cell(
                row=row_number,
                column=22,
            ).value
            shift_status = component_sheet.cell(
                row=row_number,
                column=24,
            ).value

            detail = detail_by_inspection.setdefault(
                inspection_id,
                {"angles": [], "shifts": []},
            )

            if isinstance(tilt_value, (int, float)):
                detail["angles"].append(
                    f"{component}: {float(tilt_value):+.2f}° "
                    f"({angle_status})"
                )

            if isinstance(shift_value, (int, float)):
                detail["shifts"].append(
                    f"{component}: {float(shift_value):.2f} px "
                    f"({shift_status})"
                )

        for row_number in range(2, summary_sheet.max_row + 1):
            inspection_id = summary_sheet.cell(
                row=row_number,
                column=1,
            ).value
            detail = detail_by_inspection.get(inspection_id)

            if not detail:
                continue

            summary_sheet.cell(
                row=row_number,
                column=14,
            ).value = "; ".join(detail["angles"]) or "-"
            summary_sheet.cell(
                row=row_number,
                column=15,
            ).value = "; ".join(detail["shifts"]) or "-"

    def save_excel_inspection(
        self,
        analysis,
        timestamp,
        trigger,
        raw_path,
        annotated_path,
        json_path,
        text_path,
    ):
        """Menyimpan satu baris untuk setiap komponen."""
        if not OPENPYXL_AVAILABLE:
            print(
                "[EXCEL] openpyxl belum terpasang. "
                "Jalankan: pip install openpyxl"
            )
            return None

        excel_path = self._excel_file_path()

        try:
            if os.path.exists(excel_path):
                workbook = load_workbook(excel_path)
                self._repair_existing_excel_measurements(workbook)
                self._ensure_all_component_sheets(workbook)
                self._migrate_existing_rows_to_component_sheets(
                    workbook
                )
            else:
                workbook = self._create_excel_workbook(
                    excel_path
                )

            # Pastikan daftar template selalu mengikuti isi gambar_input.
            self._refresh_template_sheet(workbook)

            summary_sheet = workbook["Ringkasan PCB"]
            component_sheet = workbook["Data Komponen"]
            component_sheets = self._ensure_all_component_sheets(
                workbook
            )

            time_data = datetime.strptime(
                timestamp,
                "%Y%m%d_%H%M%S_%f",
            )
            inspection_id = (
                f"JIG{self.active_jig}_"
                f"{timestamp}_{safe_name(trigger)}"
            )
            board_status = analysis.get(
                "decision",
                "NOT GOOD",
            )

            detected_map = {
                item["component"]: item
                for item in analysis.get("components", [])
            }

            good_components = []
            not_good_components = []
            missing_components = []
            tilted_components = []
            shifted_components = []

            # ------------------------------------------------
            # SATU BARIS UNTUK SETIAP KOMPONEN
            # ------------------------------------------------
            for component in self.expected_components:
                item = detected_map.get(component)

                if item is None:
                    template_name = "TIDAK TERDETEKSI"
                    template_condition = "HILANG"
                    mapping_text = (
                        f"{template_name} -> "
                        f"{component} (HILANG)"
                    )
                    component_status = "HILANG"
                    description = (
                        "KOMPONEN HILANG / "
                        "TIDAK TERDETEKSI"
                    )

                    missing_components.append(component)

                    component_row = [
                        inspection_id,
                        time_data.date(),
                        time_data.time(),
                        f"JIG {self.active_jig}",
                        board_status,
                        trigger,
                        template_name,
                        component,
                        template_condition,
                        mapping_text,
                        "TIDAK",
                        component_status,
                        description,
                        0.0,
                        None,
                        None,
                        None,
                        "HILANG",
                        "TIDAK",
                        None,
                        None,
                        None,
                        None,
                        "HILANG",
                        "YA",
                        os.path.abspath(raw_path),
                        os.path.abspath(annotated_path),
                    ]
                else:
                    template_name = item.get(
                        "name",
                        "-",
                    )
                    template_condition = (
                        "NOT GOOD"
                        if item.get("condition") == "NOT_GOOD"
                        else "GOOD"
                    )
                    mapping_text = (
                        f"{template_name} -> "
                        f"{component} ({template_condition})"
                    )

                    component_status = (
                        "NOT GOOD"
                        if item.get("fault", False)
                        else "GOOD"
                    )
                    description = self._component_reason(item)

                    if component_status == "GOOD":
                        good_components.append(component)
                    else:
                        not_good_components.append(component)

                    tilt = item.get("tilt_deg")
                    if tilt is not None:
                        if item.get("angle_info_only", False):
                            summary_angle_status = "INFORMASI"
                        elif item.get("angle_fault", False):
                            summary_angle_status = "NOT GOOD"
                        else:
                            summary_angle_status = "GOOD"

                        tilted_components.append(
                            f"{component}: "
                            f"{float(tilt):+.2f}° "
                            f"({summary_angle_status})"
                        )

                    shift = item.get("shift_px")
                    if shift is not None:
                        summary_shift_status = (
                            "NOT GOOD"
                            if item.get("shift_fault", False)
                            else "GOOD"
                        )
                        shifted_components.append(
                            f"{component}: "
                            f"{float(shift):.2f} px "
                            f"({summary_shift_status})"
                        )

                    if item.get("angle_info_only", False):
                        angle_status = "INFORMASI"
                    elif item.get("angle_fault", False):
                        angle_status = "NOT GOOD"
                    else:
                        angle_status = "GOOD"

                    shift_status = (
                        "NOT GOOD"
                        if item.get("shift_fault", False)
                        else "GOOD"
                    )

                    component_row = [
                        inspection_id,
                        time_data.date(),
                        time_data.time(),
                        f"JIG {self.active_jig}",
                        board_status,
                        trigger,
                        template_name,
                        component,
                        template_condition,
                        mapping_text,
                        "YA",
                        component_status,
                        description,
                        round(
                            float(
                                item.get(
                                    "confidence",
                                    0.0,
                                )
                            ) * 100.0,
                            2,
                        ),
                        None
                        if item.get(
                            "setpoint_angle_deg"
                        ) is None
                        else round(
                            float(
                                item[
                                    "setpoint_angle_deg"
                                ]
                            ),
                            2,
                        ),
                        round(
                            float(
                                item.get(
                                    "actual_angle_deg",
                                    item.get(
                                        "angle",
                                        0.0,
                                    ),
                                )
                            ),
                            2,
                        ),
                        None
                        if tilt is None
                        else round(float(tilt), 2),
                        angle_status,
                        (
                            "YA"
                            if item.get(
                                "angle_info_only",
                                False,
                            )
                            else "TIDAK"
                        ),
                        None
                        if item.get("dx_px") is None
                        else round(
                            float(item["dx_px"]),
                            2,
                        ),
                        None
                        if item.get("dy_px") is None
                        else round(
                            float(item["dy_px"]),
                            2,
                        ),
                        None
                        if shift is None
                        else round(float(shift), 2),
                        None
                        if item.get(
                            "shift_tolerance_px"
                        ) is None
                        else round(
                            float(
                                item[
                                    "shift_tolerance_px"
                                ]
                            ),
                            2,
                        ),
                        shift_status,
                        "TIDAK",
                        os.path.abspath(raw_path),
                        os.path.abspath(annotated_path),
                    ]

                component_sheet.append(component_row)
                row_number = component_sheet.max_row

                component_sheet.cell(
                    row=row_number,
                    column=2,
                ).number_format = "dd/mm/yyyy"
                component_sheet.cell(
                    row=row_number,
                    column=3,
                ).number_format = "hh:mm:ss"

                self._format_excel_measurement_cells(
                    component_sheet,
                    row_number,
                )

                self._style_excel_data_row(
                    component_sheet,
                    row_number,
                    12,
                )

                for column in (26, 27):
                    link_cell = component_sheet.cell(
                        row=row_number,
                        column=column,
                    )
                    if link_cell.value:
                        link_cell.hyperlink = str(
                            link_cell.value
                        )
                        link_cell.style = "Hyperlink"

                # Baris yang sama juga dimasukkan ke sheet komponen.
                if component not in component_sheets:
                    component_sheets[component] = (
                        self._ensure_component_sheet(
                            workbook,
                            component,
                        )
                    )

                self._append_component_row_to_sheet(
                    component_sheets[component],
                    component_row,
                )

            # Daftar NG pada ringkasan juga memasukkan komponen hilang.
            all_not_good = (
                list(not_good_components)
                + list(missing_components)
            )

            summary_row = [
                inspection_id,
                time_data.date(),
                time_data.time(),
                f"JIG {self.active_jig}",
                trigger,
                board_status,
                int(analysis.get("detected_count", 0)),
                int(analysis.get("expected_count", 0)),
                len(good_components),
                len(all_not_good),
                len(missing_components),
                ", ".join(good_components) or "-",
                ", ".join(all_not_good) or "-",
                "; ".join(tilted_components) or "-",
                "; ".join(shifted_components) or "-",
                ", ".join(missing_components) or "-",
                os.path.abspath(raw_path),
                os.path.abspath(annotated_path),
                os.path.abspath(json_path),
                os.path.abspath(text_path),
            ]

            summary_sheet.append(summary_row)
            summary_row_number = summary_sheet.max_row
            summary_sheet.cell(
                row=summary_row_number,
                column=2,
            ).number_format = "dd/mm/yyyy"
            summary_sheet.cell(
                row=summary_row_number,
                column=3,
            ).number_format = "hh:mm:ss"

            self._style_excel_data_row(
                summary_sheet,
                summary_row_number,
                6,
            )

            for column in (17, 18, 19, 20):
                link_cell = summary_sheet.cell(
                    row=summary_row_number,
                    column=column,
                )
                if link_cell.value:
                    link_cell.hyperlink = str(
                        link_cell.value
                    )
                    link_cell.style = "Hyperlink"

            summary_sheet.auto_filter.ref = (
                f"A1:{get_column_letter(summary_sheet.max_column)}"
                f"{summary_sheet.max_row}"
            )
            component_sheet.auto_filter.ref = (
                f"A1:{get_column_letter(component_sheet.max_column)}"
                f"{component_sheet.max_row}"
            )

            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
            workbook.save(excel_path)

            print(
                f"[EXCEL] Data komponen lengkap tersimpan: "
                f"{excel_path}"
            )
            return excel_path

        except PermissionError:
            print(
                "[EXCEL] File sedang dibuka di Microsoft Excel. "
                "Tutup file lalu lakukan capture kembali."
            )
            return None
        except Exception as exc:
            print(
                f"[EXCEL] Gagal menyimpan data komponen: {exc}"
            )
            return None


    def save_inspection(self, raw_frame, annotated_frame, analysis, trigger="AUTO"):
        decision = analysis["decision"].replace(" ", "_")
        status_folder = os.path.join(
            self.capture_folder,
            f"JIG_{self.active_jig}",
            safe_name(decision),
        )
        os.makedirs(status_folder, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]
        prefix = f"{timestamp}_{safe_name(trigger)}"
        raw_path = os.path.join(status_folder, prefix + "_raw.jpg")
        annotated_path = os.path.join(status_folder, prefix + "_hasil.jpg")
        json_path = os.path.join(status_folder, prefix + "_laporan.json")
        text_path = os.path.join(status_folder, prefix + "_laporan.txt")

        cv2.imwrite(raw_path, raw_frame, [cv2.IMWRITE_JPEG_QUALITY, JPEG_QUALITY])
        cv2.imwrite(annotated_path, annotated_frame, [cv2.IMWRITE_JPEG_QUALITY, JPEG_QUALITY])

        report = {
            "timestamp": datetime.now().isoformat(timespec="milliseconds"),
            "trigger": trigger,
            "jig_id": self.active_jig,
            "reference_file": os.path.basename(self.reference_file),
            "decision": analysis["decision"],
            "reference_ready": analysis["reference_ready"],
            "detected_count": analysis["detected_count"],
            "expected_count": analysis["expected_count"],
            "missing_components": analysis["missing_components"],
            "angle_tolerance_deg": ANGLE_TOLERANCE_DEG,
            "angle_good_rule": f"abs(kemiringan) < {ANGLE_TOLERANCE_DEG:.1f} derajat",
            "angle_not_good_rule": f"abs(kemiringan) >= {ANGLE_TOLERANCE_DEG:.1f} derajat",
            "min_shift_tolerance_px": MIN_SHIFT_TOLERANCE_PX,
            "components": [],
        }

        for item in analysis["components"]:
            report["components"].append(
                {
                    "component": item["component"],
                    "template": item["name"],
                    "condition_template": item["condition"],
                    "confidence": round(float(item["confidence"]), 4),
                    "setpoint_angle_deg": None
                    if item.get("setpoint_angle_deg") is None
                    else round(float(item["setpoint_angle_deg"]), 2),
                    "actual_angle_deg": round(
                        float(item.get("actual_angle_deg", item["angle"])),
                        2,
                    ),
                    "angle_absolute_deg": round(float(item["angle"]), 2),
                    "tilt_from_setpoint_deg": None
                    if item["tilt_deg"] is None
                    else round(float(item["tilt_deg"]), 2),
                    "tilt_from_reference_deg": None
                    if item["tilt_deg"] is None
                    else round(float(item["tilt_deg"]), 2),
                    "dx_px": None if item["dx_px"] is None else round(float(item["dx_px"]), 2),
                    "dy_px": None if item["dy_px"] is None else round(float(item["dy_px"]), 2),
                    "shift_px": None
                    if item["shift_px"] is None
                    else round(float(item["shift_px"]), 2),
                    "shift_percent": None
                    if item["shift_percent"] is None
                    else round(float(item["shift_percent"]), 2),
                    "status": "NOT GOOD" if item["fault"] else "GOOD",
                    "template_fault": bool(item["template_fault"]),
                    "angle_fault": bool(item["angle_fault"]),
                    "angle_ignored": False,
                    "angle_info_only": bool(item.get("angle_info_only", False)),
                    "shift_fault": bool(item["shift_fault"]),
                }
            )

        with open(json_path, "w", encoding="utf-8") as file:
            json.dump(report, file, indent=2, ensure_ascii=False)

        lines = [
            "LAPORAN INSPEKSI KOMPONEN PCB",
            "=" * 48,
            f"Waktu           : {report['timestamp']}",
            f"Pemicu capture  : {trigger}",
            f"Jig aktif       : JIG {report['jig_id']}",
            f"File referensi  : {report['reference_file']}",
            f"Keputusan       : {report['decision']}",
            f"Terdeteksi      : {report['detected_count']}/{report['expected_count']}",
            f"Komponen hilang : {', '.join(report['missing_components']) if report['missing_components'] else '-'}",
            f"Batas kemiringan: GOOD < {ANGLE_TOLERANCE_DEG:.1f} deg | NOT GOOD >= {ANGLE_TOLERANCE_DEG:.1f} deg",
            "",
            "RINCIAN KOMPONEN",
            "-" * 48,
        ]
        for item in report["components"]:
            if item["tilt_from_reference_deg"] is None:
                angle_text = "set point belum tersedia"
            else:
                info_text = (
                    " | hanya informasi untuk komponen bulat"
                    if item.get("angle_info_only", False)
                    else ""
                )
                angle_text = (
                    f"set point {item['setpoint_angle_deg']:+.2f} deg | "
                    f"aktual {item['actual_angle_deg']:+.2f} deg | "
                    f"miring dari set point "
                    f"{item['tilt_from_reference_deg']:+.2f} deg"
                    f"{info_text}"
                )
            shift_text = "N/A" if item["shift_px"] is None else f"{item['shift_px']:.2f} px"
            lines.append(
                f"{item['component']}: {item['status']} | {angle_text} | "
                f"pergeseran {shift_text} | confidence {item['confidence'] * 100:.1f}%"
            )

        if report["missing_components"]:
            lines.append("")
            lines.append("KOMPONEN TIDAK TERDETEKSI")
            lines.append("-" * 48)
            for component in report["missing_components"]:
                lines.append(
                    f"{component}: NOT GOOD | HILANG | "
                    "set point tersedia tetapi komponen tidak terdeteksi"
                )

        with open(text_path, "w", encoding="utf-8") as file:
            file.write("\n".join(lines))

        dataset_paths = None
        if analysis["decision"] in ("GOOD", "NOT GOOD"):
            dataset_paths = self.save_learning_dataset(
                raw_frame,
                annotated_frame,
                analysis,
                timestamp,
                trigger,
            )

        excel_path = self.save_excel_inspection(
            analysis=analysis,
            timestamp=timestamp,
            trigger=trigger,
            raw_path=raw_path,
            annotated_path=annotated_path,
            json_path=json_path,
            text_path=text_path,
        )

        if excel_path:
            self.last_capture_message = (
                f"CAPTURE {analysis['decision']} + DATASET + EXCEL TERSIMPAN"
            )
        else:
            self.last_capture_message = (
                f"CAPTURE {analysis['decision']} + DATASET TERSIMPAN"
            )
        self.last_capture_message_until = time.time() + 3.0
        print(f"[CAPTURE] {annotated_path}")
        print(f"[LAPORAN] {text_path}")
        if excel_path:
            print(f"[EXCEL] {excel_path}")

        result_paths = dataset_paths or {
            "annotated": annotated_path,
        }
        result_paths["excel"] = excel_path
        return result_paths

    # --------------------------------------------------------
    # TAMPILAN
    # --------------------------------------------------------
    def draw_detections(self, frame, analysis):
        output = frame.copy()
        for item in analysis["components"]:
            x, y, width, height = item["bbox"]

            if item.get("held", False):
                color = (0, 165, 255)
            elif item["fault"]:
                color = (0, 0, 255)
            else:
                color = (0, 200, 0)

            cv2.rectangle(output, (x, y), (x + width, y + height), color, 2)
            center = (int(round(item["center"][0])), int(round(item["center"][1])))
            cv2.circle(output, center, 4, color, -1)

            if self.reference and item["component"] in self.reference.get("components", {}):
                ref = self.reference["components"][item["component"]]
                ref_x = int(round(ref["center_norm"][0] * output.shape[1]))
                ref_y = int(round(ref["center_norm"][1] * output.shape[0]))
                cv2.circle(output, (ref_x, ref_y), 5, (255, 255, 0), 1)
                cv2.line(output, (ref_x, ref_y), center, color, 1)

            if item["tilt_deg"] is None:
                metric_text = f"A:{item['angle']:+.1f}deg S:N/A"
            else:
                shift_text = (
                    "N/A"
                    if item["shift_px"] is None
                    else f"{item['shift_px']:.1f}px"
                )
                info_suffix = "*" if item.get("angle_info_only", False) else ""
                metric_text = (
                    f"dA:{item['tilt_deg']:+.1f}deg{info_suffix} "
                    f"S:{shift_text}"
                )
            label = f"{item['component']} | {metric_text}"

            font = cv2.FONT_HERSHEY_SIMPLEX
            font_scale = 0.42
            thickness = 1
            text_size = cv2.getTextSize(label, font, font_scale, thickness)[0]
            text_y = max(text_size[1] + 8, y)
            cv2.rectangle(
                output,
                (x, text_y - text_size[1] - 7),
                (min(output.shape[1] - 1, x + text_size[0] + 6), text_y + 2),
                color,
                -1,
            )
            cv2.putText(
                output,
                label,
                (x + 3, text_y - 3),
                font,
                font_scale,
                (255, 255, 255),
                thickness,
                cv2.LINE_AA,
            )

        # Tandai komponen yang hilang pada posisi referensinya.
        if self.reference:
            for component in analysis.get("missing_components", []):
                missing_bbox = self._missing_component_bbox(
                    component,
                    output.shape,
                )
                if missing_bbox is None:
                    continue

                x, y, width, height = missing_bbox
                x = max(0, min(output.shape[1] - 1, int(x)))
                y = max(0, min(output.shape[0] - 1, int(y)))
                width = max(1, min(output.shape[1] - x, int(width)))
                height = max(1, min(output.shape[0] - y, int(height)))

                cv2.rectangle(
                    output,
                    (x, y),
                    (x + width, y + height),
                    (0, 0, 255),
                    3,
                )

                missing_label = f"{component} HILANG"
                text_size = cv2.getTextSize(
                    missing_label,
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.48,
                    2,
                )[0]
                label_y = max(text_size[1] + 8, y)
                cv2.rectangle(
                    output,
                    (x, label_y - text_size[1] - 7),
                    (
                        min(
                            output.shape[1] - 1,
                            x + text_size[0] + 8,
                        ),
                        label_y + 3,
                    ),
                    (0, 0, 255),
                    -1,
                )
                cv2.putText(
                    output,
                    missing_label,
                    (x + 4, label_y - 3),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.48,
                    (255, 255, 255),
                    2,
                    cv2.LINE_AA,
                )

        return output

    def draw_info_panel(self, frame, analysis, fps):
        panel = np.full((frame.shape[0], PANEL_WIDTH, 3), (28, 28, 28), dtype=np.uint8)
        line_height = 22
        y = 25

        decision = analysis["decision"]
        decision_colors = {
            "GOOD": (0, 230, 0),
            "NOT GOOD": (0, 0, 255),
            "KALIBRASI": (0, 210, 255),
            "MEMBACA": (0, 210, 255),
            "MENUNGGU PCB": (180, 180, 180),
            "TEMPLATE KOSONG": (0, 0, 255),
        }
        decision_color = decision_colors.get(decision, (255, 255, 255))

        cv2.putText(panel, "INSPEKSI KOMPONEN PCB", (12, y), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255, 255, 255), 2)
        y += 33
        cv2.putText(panel, f"STATUS: {decision}", (12, y), cv2.FONT_HERSHEY_SIMPLEX, 0.72, decision_color, 2)
        y += 30

        reference_text = "SIAP" if analysis["reference_ready"] else "BELUM ADA"
        cv2.putText(
            panel,
            f"JIG AKTIF: JIG {self.active_jig}",
            (12, y),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.52,
            (0, 220, 255),
            1,
        )
        y += line_height
        cv2.putText(
            panel,
            f"Referensi JIG {self.active_jig}: {reference_text}",
            (12, y),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.48,
            (210, 210, 210),
            1,
        )
        y += line_height
        cv2.putText(
            panel,
            f"Komponen: {analysis['detected_count']}/{analysis['expected_count']} | FPS: {fps}",
            (12, y),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.48,
            (210, 210, 210),
            1,
        )
        y += line_height
        cv2.putText(
            panel,
            f"Batas miring: GOOD < {ANGLE_TOLERANCE_DEG:.0f} deg | NG >= {ANGLE_TOLERANCE_DEG:.0f} deg",
            (12, y),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.42,
            (190, 210, 255),
            1,
        )
        y += line_height
        cv2.putText(
            panel,
            "* Sudut C1/C2 tampil sebagai informasi",
            (12, y),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.38,
            (170, 190, 220),
            1,
        )
        y += line_height

        if self.auto_capture_locked:
            capture_text = "Auto capture: TERSIMPAN (angkat PCB)"
            capture_color = (0, 200, 255)
        elif self.auto_capture_phase == "SCANNING":
            capture_text = (
                f"Scanning awal: "
                f"{self.auto_scan_remaining_seconds:.1f} detik lagi"
            )
            capture_color = (0, 220, 255)
        elif self.auto_capture_phase == "STABILIZING":
            if self.auto_stable_started_at is None:
                capture_text = (
                    f"Hasil {self.auto_capture_candidate_decision}: "
                    "menunggu stabil"
                )
            else:
                capture_text = (
                    f"Stabil {self.auto_capture_candidate_decision}: "
                    f"{self.auto_capture_remaining_seconds:.1f} detik lagi"
                )
            capture_color = (0, 220, 255)
        else:
            capture_text = "Auto scan: tunggu PCB pada jig"
            capture_color = (180, 220, 255)
        cv2.putText(panel, capture_text, (12, y), cv2.FONT_HERSHEY_SIMPLEX, 0.43, capture_color, 1)
        y += line_height + 4

        cv2.line(panel, (10, y), (PANEL_WIDTH - 10, y), (85, 85, 85), 1)
        y += 20

        missing = analysis["missing_components"]
        if missing:
            cv2.putText(panel, f"KOMPONEN HILANG ({len(missing)}):", (12, y), cv2.FONT_HERSHEY_SIMPLEX, 0.50, (0, 0, 255), 1)
            y += line_height
            missing_text = ", ".join(missing)
            # Pecah teks panjang menjadi beberapa baris.
            current_line = ""
            for token in missing_text.split(", "):
                candidate = token if not current_line else current_line + ", " + token
                if cv2.getTextSize(candidate, cv2.FONT_HERSHEY_SIMPLEX, 0.40, 1)[0][0] > PANEL_WIDTH - 25:
                    cv2.putText(panel, current_line, (12, y), cv2.FONT_HERSHEY_SIMPLEX, 0.40, (80, 120, 255), 1)
                    y += 18
                    current_line = token
                else:
                    current_line = candidate
            if current_line:
                cv2.putText(panel, current_line, (12, y), cv2.FONT_HERSHEY_SIMPLEX, 0.40, (80, 120, 255), 1)
                y += 20
        else:
            cv2.putText(panel, "KOMPONEN HILANG: TIDAK ADA", (12, y), cv2.FONT_HERSHEY_SIMPLEX, 0.48, (0, 220, 0), 1)
            y += 24

        cv2.line(panel, (10, y), (PANEL_WIDTH - 10, y), (85, 85, 85), 1)
        y += 18
        cv2.putText(panel, "RINCIAN SEMUA KOMPONEN", (12, y), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (220, 220, 220), 1)
        y += 22

        # Susun rincian berdasarkan daftar komponen yang seharusnya ada,
        # sehingga komponen GOOD, NOT GOOD, maupun HILANG semuanya tampil.
        detected_map = {
            item["component"]: item
            for item in analysis.get("components", [])
        }
        results = []
        for component in self.expected_components:
            item = detected_map.get(component)
            if item is None:
                results.append(
                    {
                        "component": component,
                        "missing": True,
                    }
                )
            else:
                shown = dict(item)
                shown["missing"] = False
                results.append(shown)

        max_scroll = max(0, len(results) - MAX_PANEL_COMPONENTS)
        self.scroll_offset = max(0, min(self.scroll_offset, max_scroll))
        visible = results[
            self.scroll_offset:self.scroll_offset + MAX_PANEL_COMPONENTS
        ]

        for item in visible:
            component_name = item["component"][:18]

            if item.get("missing", False):
                color = (0, 0, 255)
                line1 = f"[HILANG] {component_name} | tidak terdeteksi"
                line2 = "      sudut N/A | geser N/A | status NOT GOOD"
            else:
                if item.get("held", False):
                    color = (0, 165, 255)
                    status = "HOLD"
                elif item["fault"]:
                    color = (0, 0, 255)
                    status = "NG"
                else:
                    color = (0, 220, 0)
                    status = "OK"

                setpoint = item.get("setpoint_angle_deg")
                actual = item.get("actual_angle_deg", item.get("angle", 0.0))
                tilt = item.get("tilt_deg")
                shift = item.get("shift_px")

                if setpoint is None or tilt is None:
                    line1 = (
                        f"[{status}] {component_name} | "
                        f"sudut aktual {actual:+.1f} deg"
                    )
                    line2 = (
                        f"      SP N/A | geser "
                        f"{'N/A' if shift is None else f'{shift:.1f}px'}"
                    )
                else:
                    info_text = (
                        " (info)"
                        if item.get("angle_info_only", False)
                        else ""
                    )
                    line1 = (
                        f"[{status}] {component_name} | "
                        f"miring dari SP {tilt:+.1f} deg{info_text}"
                    )
                    line2 = (
                        f"      SP {setpoint:+.1f} -> aktual "
                        f"{actual:+.1f} deg | geser "
                        f"{'N/A' if shift is None else f'{shift:.1f}px'}"
                    )

            cv2.putText(
                panel,
                line1,
                (12, y),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.40,
                color,
                1,
                cv2.LINE_AA,
            )
            y += 17
            cv2.putText(
                panel,
                line2,
                (12, y),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.36,
                (190, 190, 190),
                1,
                cv2.LINE_AA,
            )
            y += 21

            if y > frame.shape[0] - 80:
                break

        if len(results) > MAX_PANEL_COMPONENTS and y < frame.shape[0] - 65:
            cv2.putText(
                panel,
                (
                    f"Scroll {self.scroll_offset + 1}-"
                    f"{min(len(results), self.scroll_offset + MAX_PANEL_COMPONENTS)}"
                    f"/{len(results)} (panah atas/bawah)"
                ),
                (12, y),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.36,
                (160, 160, 160),
                1,
            )

        footer_y = frame.shape[0] - 54
        cv2.line(panel, (10, footer_y - 12), (PANEL_WIDTH - 10, footer_y - 12), (85, 85, 85), 1)
        cv2.putText(
            panel,
            "1 JIG1 | 2 JIG2 | K kalibrasi | Z hapus ref",
            (12, footer_y),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.38,
            (205, 205, 205),
            1,
        )
        cv2.putText(
            panel,
            "Q keluar | C capture | R reload | M mute | U buka",
            (12, footer_y + 20),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.38,
            (205, 205, 205),
            1,
        )

        combined = np.hstack((frame, panel))

        if time.time() < self.last_capture_message_until:
            message = self.last_capture_message
            font = cv2.FONT_HERSHEY_SIMPLEX
            font_scale = 0.75
            thickness = 2
            size = cv2.getTextSize(message, font, font_scale, thickness)[0]
            box_x = max(10, (frame.shape[1] - size[0]) // 2)
            box_y = frame.shape[0] - 45
            cv2.rectangle(
                combined,
                (box_x - 10, box_y - size[1] - 10),
                (box_x + size[0] + 10, box_y + 8),
                (0, 0, 0),
                -1,
            )
            cv2.putText(combined, message, (box_x, box_y), font, font_scale, (0, 255, 255), thickness)

        return combined

    def reset_tracking(self):
        self.angle_smoother.clear()
        self.position_smoother.clear()
        self.tracks = {
            component: {"seen": 0, "missed": HOLD_MISSING_FRAMES + 1, "last": None}
            for component in self.expected_components
        }
        self.board_present_counter = 0
        self.reset_auto_capture_lock()

    def cleanup(self):
        self.alarm.stop_alarm()


# ============================================================
# PROGRAM UTAMA
# ============================================================
def open_camera():
    """Mencari webcam secara otomatis pada beberapa index dan backend OpenCV."""
    preferred_index = CAMERA_INDEX
    camera_indices = [preferred_index] + [i for i in range(6) if i != preferred_index]

    if sys.platform == "win32":
        backends = [
            ("DirectShow", cv2.CAP_DSHOW),
            ("Media Foundation", cv2.CAP_MSMF),
            ("Default", cv2.CAP_ANY),
        ]
    else:
        backends = [("Default", cv2.CAP_ANY)]

    print("\nMencari kamera...")

    for camera_index in camera_indices:
        for backend_name, backend in backends:
            camera = cv2.VideoCapture(camera_index, backend)

            if not camera.isOpened():
                camera.release()
                continue

            # MJPG biasanya lebih stabil untuk Logitech C920 pada 640x480/30 FPS.
            camera.set(
                cv2.CAP_PROP_FOURCC,
                cv2.VideoWriter_fourcc(*"MJPG"),
            )
            camera.set(cv2.CAP_PROP_FRAME_WIDTH, CAMERA_WIDTH)
            camera.set(cv2.CAP_PROP_FRAME_HEIGHT, CAMERA_HEIGHT)
            camera.set(cv2.CAP_PROP_FPS, CAMERA_FPS)
            camera.set(cv2.CAP_PROP_BUFFERSIZE, 1)

            # Kamera kadang perlu beberapa frame untuk mulai mengirim gambar.
            frame_ok = False
            for _ in range(15):
                success, frame = camera.read()
                if success and frame is not None and frame.size > 0:
                    frame_ok = True
                    break
                time.sleep(0.05)

            if frame_ok:
                actual_width = int(camera.get(cv2.CAP_PROP_FRAME_WIDTH))
                actual_height = int(camera.get(cv2.CAP_PROP_FRAME_HEIGHT))
                actual_fps = camera.get(cv2.CAP_PROP_FPS)
                print(
                    f"Kamera berhasil dibuka: index={camera_index}, "
                    f"backend={backend_name}, resolusi={actual_width}x{actual_height}, "
                    f"FPS={actual_fps:.1f}"
                )
                return camera

            camera.release()

    print("Kamera tidak ditemukan pada index 0 sampai 5.")
    return None


def print_startup_info(detector):
    print("\n" + "=" * 72)
    print("SISTEM INSPEKSI PCB - SCAN 5 DETIK + STABIL 5 DETIK")
    print("=" * 72)
    print(f"Folder template : {detector.template_folder}")
    print(f"Jumlah template : {len(detector.templates)}")
    print(f"Komponen unik   : {len(detector.expected_components)}")
    print(f"Jig aktif       : JIG {detector.active_jig}")
    print(f"Referensi JIG 1 : {'SIAP' if detector.has_reference(1) else 'BELUM ADA'}")
    print(f"Referensi JIG 2 : {'SIAP' if detector.has_reference(2) else 'BELUM ADA'}")
    print("\nCara kerja auto capture:")
    print("  1. Pilih jig dengan tombol 1 atau 2.")
    print("  2. Setiap jig wajib dikalibrasi sendiri memakai PCB GOOD dan tombol K.")
    print(
        f"  3. Saat minimal {AUTO_CAPTURE_MIN_DETECTED_COMPONENTS} komponen terlihat, "
        f"scanning awal berjalan {INITIAL_SCAN_SECONDS:.0f} detik."
    )
    print(
        f"  4. Setelah scanning, hasil harus stabil "
        f"{STABLE_CAPTURE_SECONDS:.0f} detik."
    )
    print("  5. Jika hasil berubah, hitungan stabil dimulai ulang.")
    print("  6. Lengkap tanpa masalah disimpan sebagai GOOD.")
    print("  7. Komponen hilang/miring/bergeser disimpan sebagai NOT GOOD.")
    print("  8. Capture hanya sekali per PCB; angkat PCB untuk membuka berikutnya.")
    print(f"Folder dataset  : {detector.dataset_folder}")
    print(f"Folder Excel    : {detector.excel_base_folder}")
    print(
        "Status Excel    : "
        + (
            "SIAP - DATA KOMPONEN LENGKAP"
            if OPENPYXL_AVAILABLE
            else "OPENPYXL BELUM TERPASANG"
        )
    )
    print("Format Excel    : 1 file per jig + 1 sheet per komponen")
    print(f"Batas kemiringan: GOOD < {ANGLE_TOLERANCE_DEG:.0f} derajat | NOT GOOD >= {ANGLE_TOLERANCE_DEG:.0f} derajat")
    print("Catatan          : sudut semua komponen ditampilkan; sudut C1/C2 hanya informasi.")
    print("\nKontrol:")
    print("  1 = gunakan profil JIG 1")
    print("  2 = gunakan profil JIG 2")
    print("  Q = keluar")
    print("  C = capture manual")
    print("  K = kalibrasi/simpan referensi untuk jig aktif")
    print("  Z = hapus referensi hanya untuk jig aktif")
    print("  R = reload template")
    print("  M = mute/unmute alarm")
    print("  U = buka kunci auto capture")
    print("  Panah atas/bawah = scroll rincian")
    print("=" * 72)

    if detector.templates:
        for item in detector.templates:
            print(
                f"  - {item['name']} -> {item['component']} "
                f"({'NOT GOOD' if item['condition'] == 'NOT_GOOD' else 'GOOD'})"
            )
    else:
        print("[PERINGATAN] Folder gambar_input belum berisi template.")


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    detector = ComponentDetector(base_dir=base_dir, template_folder="gambar_input")
    camera = open_camera()

    if camera is None:
        print("Tidak dapat membuka webcam. Tutup Camera/Zoom/OBS, periksa izin kamera Windows, lalu jalankan kembali.")
        return

    window_name = "Sistem Inspeksi Komponen PCB"
    cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
    cv2.resizeWindow(window_name, CAMERA_WIDTH + PANEL_WIDTH, CAMERA_HEIGHT + 80)
    print_startup_info(detector)

    fps_counter = 0
    fps_value = 0
    fps_start = time.time()

    last_raw_frame = None
    last_combined_frame = None
    last_analysis = None
    last_detections = []

    try:
        while True:
            success, frame = camera.read()
            if not success:
                print("Gagal membaca frame dari kamera.")
                break

            if FLIP_CAMERA:
                frame = cv2.flip(frame, 1)

            detections = detector.detect_components(frame)
            analysis = detector.analyze_board(detections, frame.shape)
            auto_action = detector.update_auto_capture_state(analysis)

            fps_counter += 1
            elapsed = time.time() - fps_start
            if elapsed >= 1.0:
                fps_value = int(round(fps_counter / elapsed))
                fps_counter = 0
                fps_start = time.time()

            annotated = detector.draw_detections(frame, analysis)
            combined = detector.draw_info_panel(annotated, analysis, fps_value)

            if auto_action == "CAPTURE":
                detector.save_inspection(frame, combined, analysis, trigger="AUTO")
                # Gambar ulang supaya pesan capture langsung terlihat.
                combined = detector.draw_info_panel(annotated, analysis, fps_value)

            cv2.imshow(window_name, combined)

            last_raw_frame = frame.copy()
            last_combined_frame = combined.copy()
            last_analysis = analysis
            last_detections = detections

            key_ex = cv2.waitKeyEx(1)
            key = key_ex & 0xFF

            if key == ord("q"):
                print("Keluar dari program...")
                break

            if key == ord("1"):
                detector.switch_jig(1)

            elif key == ord("2"):
                detector.switch_jig(2)

            elif key == ord("c"):
                if last_analysis is None:
                    print("Belum ada frame yang dapat disimpan.")
                else:
                    detector.save_inspection(
                        last_raw_frame,
                        last_combined_frame,
                        last_analysis,
                        trigger="MANUAL",
                    )

            elif key == ord("k"):
                if last_analysis and last_analysis["complete_live"]:
                    if detector.save_reference(last_detections, last_raw_frame.shape):
                        detector.reset_tracking()
                        print(f"Referensi JIG {detector.active_jig} diperbarui. Tahan PCB sampai pembacaan stabil kembali.")
                else:
                    missing = last_analysis["missing_components"] if last_analysis else []
                    print(
                        f"Referensi JIG {detector.active_jig} belum dapat disimpan. Komponen belum lengkap:",
                        ", ".join(missing) or "belum stabil",
                    )

            elif key == ord("z"):
                detector.delete_reference()

            elif key == ord("r"):
                print("Reload template...")
                detector.load_templates()
                detector.load_reference()
                print(f"Template dimuat: {len(detector.templates)}")

            elif key == ord("m"):
                muted = detector.alarm.toggle_mute()
                print("Alarm MUTED" if muted else "Alarm UNMUTED")

            elif key == ord("u"):
                detector.unlock_auto_capture()

            # Kode panah dapat berbeda antara Windows dan Linux.
            elif key_ex in (82, 2490368, 65362):
                detector.scroll_offset = max(0, detector.scroll_offset - 1)
            elif key_ex in (84, 2621440, 65364):
                detector.scroll_offset += 1

    except KeyboardInterrupt:
        print("Program dihentikan oleh pengguna.")
    finally:
        detector.cleanup()
        camera.release()
        cv2.destroyAllWindows()
        print("Program selesai.")


if __name__ == "__main__":
    main()